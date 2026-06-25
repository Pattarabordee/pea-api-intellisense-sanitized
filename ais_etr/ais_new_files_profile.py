from __future__ import annotations

from collections import Counter
import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_OUTPUT_DIR = Path("runtime/analysis")

CATALOG_COLUMNS = [
    "source_id",
    "file_name",
    "path",
    "size_bytes",
    "last_modified",
    "source_owner",
    "role",
    "status",
    "notes",
]

JOIN_AUDIT_COLUMNS = [
    "join_candidate",
    "alarm_key_column",
    "mapping_key_column",
    "alarm_rows_matched",
    "alarm_unique_keys_matched",
    "alarm_unique_keys_total",
    "mapping_unique_keys",
    "mapping_duplicate_keys",
    "recommendation",
]


@dataclass
class AlarmProfile:
    source_path: Path
    rows: int
    columns: list[str]
    non_empty_counts: dict[str, int]
    sitecode_row_counts: Counter[str]
    first_occurrence_min: datetime | None
    first_occurrence_max: datetime | None
    missing_counts: dict[str, int]
    jobid_missing: int
    ticketid_missing: int
    duration_rows: int
    negative_duration_rows: int
    duration_bands: dict[str, int]
    alarmname_counts: dict[str, int]
    year_counts: dict[str, int]
    severity_counts: dict[str, int]


@dataclass
class MappingProfile:
    source_path: Path
    sheet_name: str
    rows: int
    headers: list[str]
    non_empty_counts: dict[str, int]
    duplicate_key_counts: dict[str, int]
    key_sets: dict[str, set[str]]
    key_duplicate_counts: dict[str, int]
    province_counts: dict[str, int]
    company_counts: dict[str, int]
    group_counts: dict[str, int]


def build_ais_new_files_profile(
    ac_alarm_csv: str | Path = "AC MAIN FAIL.csv",
    meter_mapping_xlsx: str | Path | None = None,
    legacy_alarm_xlsx: str | Path | None = "NE_FAC_AC MAIN FAIL.xlsx",
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
) -> dict[str, Any]:
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    alarm_path = Path(ac_alarm_csv)
    mapping_path = Path(meter_mapping_xlsx) if meter_mapping_xlsx else _find_meter_mapping_file()
    legacy_path = Path(legacy_alarm_xlsx) if legacy_alarm_xlsx else None

    alarm_profile = profile_ac_main_fail_csv(alarm_path)
    mapping_profile = profile_meter_mapping_xlsx(mapping_path)
    join_rows = build_join_audit(alarm_profile, mapping_profile)

    catalog_path = output_root / "ais_source_catalog.csv"
    alarm_dict_path = output_root / "ac_main_fail_column_dictionary.csv"
    alarm_profile_path = output_root / "ac_main_fail_profile.csv"
    mapping_profile_path = output_root / "meter_mapping_profile.csv"
    mapping_key_path = output_root / "meter_mapping_key_coverage.csv"
    join_audit_path = output_root / "ais_new_files_join_audit.csv"
    report_path = output_root / "ais_new_files_readiness_report.md"

    _write_csv(catalog_path, CATALOG_COLUMNS, _catalog_rows(alarm_path, mapping_path, legacy_path))
    _write_csv(alarm_dict_path, _alarm_dictionary_columns(), _alarm_column_dictionary_rows(alarm_profile))
    _write_metric_csv(alarm_profile_path, _alarm_metric_rows(alarm_profile))
    _write_metric_csv(mapping_profile_path, _mapping_metric_rows(mapping_profile))
    _write_metric_csv(mapping_key_path, _mapping_key_metric_rows(mapping_profile))
    _write_csv(join_audit_path, JOIN_AUDIT_COLUMNS, join_rows)
    report_path.write_text(
        _render_markdown_report(alarm_profile, mapping_profile, join_rows, catalog_path),
        encoding="utf-8-sig",
    )

    return {
        "output_dir": str(output_root),
        "catalog": str(catalog_path),
        "alarm_column_dictionary": str(alarm_dict_path),
        "alarm_profile": str(alarm_profile_path),
        "mapping_profile": str(mapping_profile_path),
        "mapping_key_coverage": str(mapping_key_path),
        "join_audit": str(join_audit_path),
        "report": str(report_path),
        "alarm_rows": alarm_profile.rows,
        "alarm_columns": len(alarm_profile.columns),
        "mapping_rows": mapping_profile.rows,
        "mapping_rows_with_peano": mapping_profile.non_empty_counts.get("pea_meter", 0),
        "mapping_rows_with_transformer_peano": mapping_profile.non_empty_counts.get("transformer_peano", 0),
        "best_join": join_rows[0] if join_rows else None,
        "phase1_sustained_candidate_rows": _phase1_sustained_candidate_rows(alarm_profile),
        "phase1_reject_rows": _phase1_reject_rows(alarm_profile),
        "truth_status": "candidate_not_activated",
    }


def profile_ac_main_fail_csv(path: str | Path) -> AlarmProfile:
    source = Path(path)
    non_empty: Counter[str] = Counter()
    sitecode_row_counts: Counter[str] = Counter()
    missing: Counter[str] = Counter()
    alarmname_counts: Counter[str] = Counter()
    year_counts: Counter[str] = Counter()
    severity_counts: Counter[str] = Counter()
    rows = 0
    first_min: datetime | None = None
    first_max: datetime | None = None
    jobid_missing = 0
    ticketid_missing = 0
    duration_rows = 0
    negative_duration_rows = 0
    duration_bands: Counter[str] = Counter()

    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        for row in reader:
            rows += 1
            for column in columns:
                if _text(row.get(column)):
                    non_empty[column] += 1

            sitecode = _norm(row.get("Sitecode"))
            if sitecode:
                sitecode_row_counts[sitecode] += 1
            else:
                missing["Sitecode"] += 1

            first_occurrence = _parse_dt(row.get("Firstoccurrence"))
            clear_time = _parse_dt(row.get("Cleartime"))
            if not first_occurrence:
                missing["Firstoccurrence"] += 1
            if not clear_time:
                missing["Cleartime"] += 1
            if not _text(row.get("Jobid")):
                jobid_missing += 1
            if not _text(row.get("Ticketid")):
                ticketid_missing += 1

            if first_occurrence:
                first_min = first_occurrence if first_min is None or first_occurrence < first_min else first_min
                first_max = first_occurrence if first_max is None or first_occurrence > first_max else first_max
            if first_occurrence and clear_time:
                duration = (clear_time - first_occurrence).total_seconds() / 60
                duration_rows += 1
                if duration < 0:
                    negative_duration_rows += 1
                duration_bands[_duration_band(duration)] += 1

            _count_if_present(alarmname_counts, row.get("Alarmname"))
            _count_if_present(year_counts, row.get("Year of Firstoccurrence"))
            _count_if_present(severity_counts, row.get("Severity"))

    return AlarmProfile(
        source_path=source,
        rows=rows,
        columns=columns,
        non_empty_counts=dict(non_empty),
        sitecode_row_counts=sitecode_row_counts,
        first_occurrence_min=first_min,
        first_occurrence_max=first_max,
        missing_counts=dict(missing),
        jobid_missing=jobid_missing,
        ticketid_missing=ticketid_missing,
        duration_rows=duration_rows,
        negative_duration_rows=negative_duration_rows,
        duration_bands=dict(duration_bands),
        alarmname_counts=dict(alarmname_counts.most_common(12)),
        year_counts=dict(year_counts.most_common()),
        severity_counts=dict(severity_counts.most_common()),
    )


def profile_meter_mapping_xlsx(path: str | Path, sheet_name: str = "Joined") -> MappingProfile:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        headers = [_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        indices = _mapping_indices(headers)
        rows = 0
        non_empty: Counter[str] = Counter()
        key_values: dict[str, Counter[str]] = {
            "site_code": Counter(),
            "location_id": Counter(),
            "meter": Counter(),
            "ca_format": Counter(),
            "ca": Counter(),
            "pea_meter": Counter(),
            "transformer_peano": Counter(),
        }
        province_counts: Counter[str] = Counter()
        company_counts: Counter[str] = Counter()
        group_counts: Counter[str] = Counter()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            rows += 1
            for key, index in indices.items():
                value = _cell(row, index)
                if value:
                    non_empty[key] += 1
            for key in key_values:
                value = _norm(_cell(row, indices[key]))
                if value:
                    key_values[key][value] += 1
            _count_if_present(province_counts, _cell(row, indices["province"]))
            _count_if_present(company_counts, _cell(row, indices["com"]))
            _count_if_present(company_counts, _cell(row, indices["company"]))
            _count_if_present(group_counts, _cell(row, indices["group"]))

        duplicate_key_counts = {
            key: sum(1 for count in values.values() if count > 1)
            for key, values in key_values.items()
        }
        return MappingProfile(
            source_path=source,
            sheet_name=worksheet.title,
            rows=rows,
            headers=headers,
            non_empty_counts=dict(non_empty),
            duplicate_key_counts=duplicate_key_counts,
            key_sets={key: set(values) for key, values in key_values.items()},
            key_duplicate_counts=duplicate_key_counts,
            province_counts=dict(province_counts.most_common(20)),
            company_counts=dict(company_counts.most_common(20)),
            group_counts=dict(group_counts.most_common(20)),
        )
    finally:
        workbook.close()


def build_join_audit(alarm: AlarmProfile, mapping: MappingProfile) -> list[dict[str, Any]]:
    alarm_total_unique = len(alarm.sitecode_row_counts)
    rows = []
    candidates = [
        ("sitecode_to_site_code", "SITE Code", "site_code"),
        ("sitecode_to_location_id", "Location ID", "location_id"),
        ("sitecode_to_meter", "Meter", "meter"),
        ("sitecode_to_ca_format", "CA_FORMAT", "ca_format"),
        ("sitecode_to_ca", "CA", "ca"),
        ("sitecode_to_pea_meter", "PEA meter", "pea_meter"),
        ("sitecode_to_transformer_peano", "Transformer PEANO", "transformer_peano"),
    ]
    for name, label, key in candidates:
        mapping_keys = mapping.key_sets.get(key, set())
        matched_keys = set(alarm.sitecode_row_counts) & mapping_keys
        matched_rows = sum(alarm.sitecode_row_counts[value] for value in matched_keys)
        duplicate_keys = mapping.key_duplicate_counts.get(key, 0)
        rows.append(
            {
                "join_candidate": name,
                "alarm_key_column": "Sitecode",
                "mapping_key_column": label,
                "alarm_rows_matched": matched_rows,
                "alarm_unique_keys_matched": len(matched_keys),
                "alarm_unique_keys_total": alarm_total_unique,
                "mapping_unique_keys": len(mapping_keys),
                "mapping_duplicate_keys": duplicate_keys,
                "recommendation": _join_recommendation(matched_rows, alarm.rows, duplicate_keys, key),
            }
        )
    return sorted(rows, key=lambda row: (int(row["alarm_rows_matched"]), int(row["alarm_unique_keys_matched"])), reverse=True)


def _catalog_rows(alarm_path: Path, mapping_path: Path, legacy_path: Path | None) -> list[dict[str, str]]:
    rows = [
        _catalog_row(
            "ais_ac_main_fail_csv",
            alarm_path,
            "AIS",
            "raw_alarm",
            "raw_candidate",
            "Detailed AC MAIN FAIL alarm export; candidate AIS site power-impact truth.",
        ),
        _catalog_row(
            "ais_meter_id_ne_mapping",
            mapping_path,
            "AIS",
            "site_meter_mapping",
            "mapping_candidate",
            "Candidate SITE/PEANO/lat-long mapping; direct Sitecode coverage is weak.",
        ),
    ]
    if legacy_path:
        rows.append(
            _catalog_row(
                "ais_ne_fac_ac_main_fail_legacy",
                legacy_path,
                "AIS",
                "legacy_alarm_workbook",
                "needs_owner_clarification",
                "Earlier workbook with raw and pivot tabs; keep as reference until CSV supersedes it.",
            )
        )
    return rows


def _catalog_row(source_id: str, path: Path, owner: str, role: str, status: str, notes: str) -> dict[str, str]:
    info = path.stat()
    return {
        "source_id": source_id,
        "file_name": path.name,
        "path": str(path.resolve()),
        "size_bytes": str(info.st_size),
        "last_modified": datetime.fromtimestamp(info.st_mtime).isoformat(sep=" ", timespec="seconds"),
        "source_owner": owner,
        "role": role,
        "status": status,
        "notes": notes,
    }


def _alarm_column_dictionary_rows(profile: AlarmProfile) -> list[dict[str, Any]]:
    rows = []
    for column in profile.columns:
        rows.append(
            {
                "column_name": column,
                "semantic_category": _column_category(column),
                "proposed_use": _proposed_column_use(column),
                "non_empty_rows": profile.non_empty_counts.get(column, 0),
                "missing_rows": profile.rows - profile.non_empty_counts.get(column, 0),
                "required_for_phase1": "yes" if column in {"Firstoccurrence", "Cleartime", "Sitecode"} else "no",
            }
        )
    return rows


def _alarm_dictionary_columns() -> list[str]:
    return [
        "column_name",
        "semantic_category",
        "proposed_use",
        "non_empty_rows",
        "missing_rows",
        "required_for_phase1",
    ]


def _alarm_metric_rows(profile: AlarmProfile) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = [
        ("source_file", profile.source_path.name),
        ("rows", profile.rows),
        ("columns", len(profile.columns)),
        ("firstoccurrence_min", _fmt_dt(profile.first_occurrence_min)),
        ("firstoccurrence_max", _fmt_dt(profile.first_occurrence_max)),
        ("unique_sitecode", len(profile.sitecode_row_counts)),
        ("missing_firstoccurrence", profile.missing_counts.get("Firstoccurrence", 0)),
        ("missing_cleartime", profile.missing_counts.get("Cleartime", 0)),
        ("missing_sitecode", profile.missing_counts.get("Sitecode", 0)),
        ("jobid_missing", profile.jobid_missing),
        ("ticketid_missing", profile.ticketid_missing),
        ("duration_rows", profile.duration_rows),
        ("negative_duration_rows", profile.negative_duration_rows),
    ]
    rows.extend((f"duration_band_{key}", value) for key, value in sorted(profile.duration_bands.items()))
    rows.extend((f"alarmname_top_{key}", value) for key, value in profile.alarmname_counts.items())
    rows.extend((f"year_{key}", value) for key, value in profile.year_counts.items())
    rows.extend((f"severity_{key}", value) for key, value in profile.severity_counts.items())
    return rows


def _mapping_metric_rows(profile: MappingProfile) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = [
        ("source_file", profile.source_path.name),
        ("sheet_name", profile.sheet_name),
        ("rows", profile.rows),
        ("columns", len(profile.headers)),
        ("rows_with_pea_meter", profile.non_empty_counts.get("pea_meter", 0)),
        ("rows_with_transformer_peano", profile.non_empty_counts.get("transformer_peano", 0)),
        ("rows_with_lat", profile.non_empty_counts.get("lat", 0)),
        ("rows_with_long", profile.non_empty_counts.get("long", 0)),
        ("rows_with_site_code", profile.non_empty_counts.get("site_code", 0)),
        ("duplicate_location_id_values", profile.duplicate_key_counts.get("location_id", 0)),
        ("duplicate_site_code_values", profile.duplicate_key_counts.get("site_code", 0)),
        ("duplicate_pea_meter_values", profile.duplicate_key_counts.get("pea_meter", 0)),
    ]
    rows.extend((f"province_top_{key}", value) for key, value in profile.province_counts.items())
    rows.extend((f"company_{key}", value) for key, value in profile.company_counts.items())
    rows.extend((f"group_{key}", value) for key, value in profile.group_counts.items())
    return rows


def _mapping_key_metric_rows(profile: MappingProfile) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    for key in sorted(profile.key_sets):
        rows.append((f"{key}_non_empty_rows", profile.non_empty_counts.get(key, 0)))
        rows.append((f"{key}_unique_values", len(profile.key_sets.get(key, set()))))
        rows.append((f"{key}_duplicate_values", profile.duplicate_key_counts.get(key, 0)))
    return rows


def _render_markdown_report(
    alarm: AlarmProfile,
    mapping: MappingProfile,
    join_rows: list[dict[str, Any]],
    catalog_path: Path,
) -> str:
    best = join_rows[0] if join_rows else {}
    sustained = _phase1_sustained_candidate_rows(alarm)
    review = _sum_bands(alarm.duration_bands, {"<=1_min", ">1_to_5_min"})
    reject = _phase1_reject_rows(alarm)
    lines = [
        "# AIS New Files Profiling And Import Readiness",
        "",
        "This report profiles the two new AIS files without moving source files or activating them as production truth.",
        "",
        "## Data Catalog",
        "",
        f"- Catalog file: `{catalog_path.as_posix()}`",
        "- Source organization mode: manifest-first, non-destructive.",
        "- Raw source files remain in their current workspace paths.",
        "",
        "## AC MAIN FAIL.csv Profile",
        "",
        f"- Rows: {alarm.rows}",
        f"- Columns: {len(alarm.columns)}",
        f"- First occurrence range: {_fmt_dt(alarm.first_occurrence_min)} to {_fmt_dt(alarm.first_occurrence_max)}",
        f"- Unique Sitecode values: {len(alarm.sitecode_row_counts)}",
        f"- Missing Cleartime rows: {alarm.missing_counts.get('Cleartime', 0)}",
        f"- Negative duration rows: {alarm.negative_duration_rows}",
        f"- Reject rows by Phase 1 policy: {reject}",
        f"- Review-only rows (`<=5 min`): {review}",
        f"- Sustained candidate rows (`>5 and <=1440 min`): {sustained}",
        f"- Missing Jobid rows: {alarm.jobid_missing}",
        f"- Missing Ticketid rows: {alarm.ticketid_missing}",
        "",
        "## Meter Mapping Profile",
        "",
        f"- Rows in `Joined`: {mapping.rows}",
        f"- Rows with PEA meter: {mapping.non_empty_counts.get('pea_meter', 0)}",
        f"- Rows with transformer PEANO: {mapping.non_empty_counts.get('transformer_peano', 0)}",
        f"- Rows with lat/long: {mapping.non_empty_counts.get('lat', 0)} / {mapping.non_empty_counts.get('long', 0)}",
        f"- Rows with SITE Code: {mapping.non_empty_counts.get('site_code', 0)}",
        f"- Duplicate Location ID values: {mapping.duplicate_key_counts.get('location_id', 0)}",
        f"- Duplicate SITE Code values: {mapping.duplicate_key_counts.get('site_code', 0)}",
        f"- Duplicate PEA meter values: {mapping.duplicate_key_counts.get('pea_meter', 0)}",
        "",
        "## Join-Key Readiness",
        "",
        "| Candidate | Alarm rows matched | Unique keys matched | Recommendation |",
        "| --- | ---: | ---: | --- |",
    ]
    for row in join_rows:
        lines.append(
            f"| {row['join_candidate']} | {row['alarm_rows_matched']} | "
            f"{row['alarm_unique_keys_matched']} | {row['recommendation']} |"
        )
    lines.extend(
        [
            "",
            "Current best exact join:",
            f"`{best.get('join_candidate', 'none')}` with {best.get('alarm_rows_matched', 0)} alarm rows matched.",
            "",
            "Decision: direct join is not strong enough to activate AIS alarm truth. Ask AIS to confirm or provide the mapping key between alarm `Sitecode` and PEANO/site registry.",
            "",
            "## Locked AIS Decisions",
            "",
            "- AIS will provide the `Sitecode` to PEANO mapping later.",
            "- Current AIS mapping understanding: `1 PEA meter = 1 site`.",
            "- AIS-side versus PEA-side fault source is not reliably classified in the alarm export.",
            "- `Cleartime` missing rows are rejected in Phase 1.",
            "- Negative duration and `>24h` duration rows are rejected in Phase 1.",
            "- `<=5 minutes` is review-only; `>5 and <=1440 minutes` is sustained-eligible.",
            "- Flapping/duplicate merge is deferred to Phase 2; Phase 1 keeps one alarm row as one candidate interval.",
            "",
            "## Provisional Truth Logic (Not Activated)",
            "",
            "- `outage_start_time = Firstoccurrence`",
            "- `power_restore_time = Cleartime`",
            "- `actual_restoration_minutes = Cleartime - Firstoccurrence`",
            "- `<=5 minutes` stays review-only.",
            "- `>5 and <=1440 minutes` is a sustained candidate.",
            "- Missing `Cleartime`, negative duration, and `>24h` duration rows are rejected.",
            "- Because AIS-side versus PEA-side fault source is not classified, only rows that later match Webex/PEA context can be used for the PEA ETR model gate.",
            "- Phase 1 does not merge flapping/duplicate alarms; one alarm row remains one candidate interval.",
            "",
            "## Next Actions",
            "",
            "1. Ask AIS for the authoritative join key from `Sitecode` to PEANO or a complete site registry.",
            "2. Keep this CSV as a candidate raw alarm source, not canonical truth yet.",
            "3. Re-run this profile after AIS supplies an improved mapping.",
            "4. Only then build the AIS truth importer and shadow match evaluation.",
            "",
            "## Safety Notes",
            "",
            "- This report contains aggregate counts only.",
            "- It does not include PEANO lists, ticket id lists, long alarm descriptions, Webex identifiers, room identifiers, or credential values.",
        ]
    )
    return "\n".join(lines) + "\n"


def _mapping_indices(headers: list[str]) -> dict[str, int | None]:
    return {
        "id": _find_index(headers, ("ID",), 1),
        "meter": _find_index(headers, ("Meter",), 3),
        "ca_format": _find_index(headers, ("CA_FORMAT",), 4),
        "com": _find_index(headers, ("Com",), 5),
        "location_id": _find_index(headers, ("Location ID", "LocationID"), 6),
        "province": _find_index(headers, ("Province",), 7),
        "group": _find_index(headers, ("Group",), 8),
        "pea_meter": _find_index(headers, ("PEA Meter", "PEANO", "Meter PEA"), 9),
        "ca": _find_index(headers, ("CA",), 10),
        "transformer_peano": _find_index(headers, ("Transformer PEANO", "TX PEANO"), 11),
        "lat": _find_index(headers, ("LAT", "Latitude"), 24),
        "long": _find_index(headers, ("LONG", "Longitude", "LON", "Lng"), 25),
        "site_code": _find_index(headers, ("SITE Code", "Sitecode", "Site Code"), 27),
        "company": _find_index(headers, ("Companay", "Company"), 28),
    }


def _find_index(headers: list[str], aliases: Iterable[str], fallback_1_based: int) -> int | None:
    normalized = {_norm(header): index for index, header in enumerate(headers)}
    for alias in aliases:
        found = normalized.get(_norm(alias))
        if found is not None:
            return found
    fallback = fallback_1_based - 1
    return fallback if 0 <= fallback < len(headers) else None


def _cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _text(row[index])


def _find_meter_mapping_file() -> Path:
    matches = sorted(Path(".").glob("Meter_ID_NE For PEA_*LatLong_R01 1.xlsx"))
    if not matches:
        raise FileNotFoundError("Meter mapping workbook not found: Meter_ID_NE For PEA_*LatLong_R01 1.xlsx")
    return matches[0]


def _write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows({column: row.get(column, "") for column in columns} for row in rows)


def _write_metric_csv(path: Path, rows: list[tuple[str, Any]]) -> None:
    _write_csv(path, ["metric", "value"], [{"metric": key, "value": value} for key, value in rows])


def _parse_dt(value: Any) -> datetime | None:
    text = _text(value)
    if not text:
        return None
    for fmt in (
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _duration_band(minutes: float) -> str:
    if minutes < 0:
        return "invalid_negative"
    if minutes <= 1:
        return "<=1_min"
    if minutes <= 5:
        return ">1_to_5_min"
    if minutes <= 60:
        return ">5_to_60_min"
    if minutes <= 180:
        return ">60_to_180_min"
    if minutes <= 1440:
        return ">180_to_1440_min"
    return ">1440_min"


def _column_category(column: str) -> str:
    lowered = column.lower()
    if any(token in lowered for token in ("time", "occurrence", "clear", "collect", "insert")):
        return "time"
    if any(token in lowered for token in ("alarm", "alert", "description", "probablecause")):
        return "alarm"
    if any(token in lowered for token in ("site", "node", "device", "ne", "moname")):
        return "network_object"
    if any(token in lowered for token in ("job", "ticket", "tt")):
        return "ticket"
    if any(token in lowered for token in ("province", "county", "township", "location")):
        return "location"
    if any(token in lowered for token in ("flag", "status", "severity", "acknowledged")):
        return "status"
    return "other"


def _proposed_column_use(column: str) -> str:
    uses = {
        "Firstoccurrence": "candidate outage_start_time",
        "Cleartime": "candidate power_restore_time",
        "Clearalarmfirstreceivetime": "restore timestamp validation",
        "Sitecode": "primary alarm-side join candidate",
        "Alarmname": "alarm type filter/feature",
        "Jobid": "external incident join candidate if populated",
        "Ticketid": "external incident join candidate if populated",
        "Flappingcount": "phase 2 flapping analysis",
        "Severity": "open/cleared state quality check",
        "Outageflag": "AIS-side outage classification candidate",
        "Sitepoweroff": "AIS-side power state flag candidate",
        "Sitedownflag": "AIS-side site-down classification candidate",
    }
    return uses.get(column, "")


def _join_recommendation(matched_rows: int, total_rows: int, duplicate_keys: int, key: str) -> str:
    if matched_rows == 0:
        return "do_not_use_no_exact_match"
    coverage = matched_rows / max(1, total_rows)
    if key == "site_code" and coverage < 0.8:
        return "weak_do_not_use_as_primary"
    if coverage < 0.8:
        return "audit_only_low_coverage"
    if duplicate_keys:
        return "needs_duplicate_resolution"
    return "candidate_primary_join"


def _count_if_present(counter: Counter[str], value: Any) -> None:
    text = _text(value)
    if text:
        counter[text] += 1


def _sum_bands(bands: dict[str, int], keys: set[str]) -> int:
    return sum(int(bands.get(key, 0)) for key in keys)


def _phase1_sustained_candidate_rows(profile: AlarmProfile) -> int:
    return _sum_bands(profile.duration_bands, {">5_to_60_min", ">60_to_180_min", ">180_to_1440_min"})


def _phase1_reject_rows(profile: AlarmProfile) -> int:
    return (
        int(profile.missing_counts.get("Firstoccurrence", 0))
        + int(profile.missing_counts.get("Cleartime", 0))
        + int(profile.negative_duration_rows)
        + int(profile.duration_bands.get(">1440_min", 0))
    )


def _norm(value: Any) -> str:
    return _text(value).upper().replace(" ", "")


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _fmt_dt(value: datetime | None) -> str:
    return value.isoformat(sep=" ") if value else ""
