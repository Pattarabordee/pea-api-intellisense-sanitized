from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
import re
from typing import Any, Iterable

from .ais_truth import AIS_TRUTH_COLUMNS, AIS_TRUTH_DEFINITION, AIS_TRUTH_SOURCE_DEFAULT, AIS_TRUTH_TARGET

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - only exercised if optional spreadsheet support is absent
    load_workbook = None  # type: ignore[assignment]


DEFAULT_ADD_FIELD_SOURCE = Path("AC_MAIN_FAIL_add_field.xlsx")
DEFAULT_ADD_FIELD_SHEET = "AC MAIN FAIL"
DEFAULT_TRUTH_OUTPUT = Path("runtime/ais_truth_latest_candidate.csv")
DEFAULT_REVIEW_OUTPUT = Path("runtime/ais_truth_review_le_5min.csv")
DEFAULT_REJECTS_OUTPUT = Path("runtime/ais_truth_rejects_add_field.csv")
DEFAULT_AUDIT_OUTPUT = Path("runtime/ais_truth_join_audit.csv")
DEFAULT_REPORT_OUTPUT = Path("runtime/analysis/ais_add_field_truth_import_report.md")

SOURCE_LABEL = "AIS_AC_MAIN_FAIL_ADD_FIELD"

AUDIT_COLUMNS = (
    "source_row_number",
    "job_id",
    "location_id",
    "sitecode",
    "outage_start_time",
    "power_restore_time",
    "actual_restoration_minutes",
    "source_duration_minutes",
    "duration_delta_minutes",
    "duration_consistency",
    "mapping_status",
    "mapped_peano_count",
    "truth_quality",
    "main_cause",
    "subcause1",
    "subcause2",
    "cause_category",
    "alarm_type",
)

_EMPTY_VALUES = {"", "-", "nan", "none", "null", "nat"}


@dataclass(frozen=True)
class MappedSite:
    location_id: str
    peanos: tuple[str, ...]
    site_codes: tuple[str, ...]


def import_ais_add_field_truth(
    source: str | Path = DEFAULT_ADD_FIELD_SOURCE,
    meter_mapping: str | Path | None = None,
    output_csv: str | Path = DEFAULT_TRUTH_OUTPUT,
    review_csv: str | Path = DEFAULT_REVIEW_OUTPUT,
    rejects_csv: str | Path = DEFAULT_REJECTS_OUTPUT,
    audit_csv: str | Path = DEFAULT_AUDIT_OUTPUT,
    report_markdown: str | Path | None = DEFAULT_REPORT_OUTPUT,
    *,
    sheet: str | int | None = DEFAULT_ADD_FIELD_SHEET,
    date_order: str = "mdy",
) -> dict[str, Any]:
    """Convert AIS add-field AC mains alarms into canonical AIS truth rows.

    The importer uses `Location ID` as the join key to the AIS meter mapping workbook.
    Only rows with exactly one mapped meter PEANO can become eligible `OK` truth.
    Rows with <=5 minute durations are preserved as REVIEW_SHORT and excluded from
    default shadow matching/model gates.
    """

    source_path = Path(source)
    mapping_path = Path(meter_mapping) if meter_mapping else _find_default_mapping(source_path.parent)
    mapping = _load_meter_mapping(mapping_path)
    raw_rows, detected_columns = _read_alarm_rows(source_path, sheet=sheet)

    canonical_rows: list[dict[str, str]] = []
    review_rows: list[dict[str, str]] = []
    reject_rows: list[dict[str, str]] = []
    audit_rows: list[dict[str, str]] = []
    quality_counts: Counter[str] = Counter()
    mapping_counts: Counter[str] = Counter()
    cause_counts: Counter[str] = Counter()
    duration_consistency_counts: Counter[str] = Counter()

    for source_row_number, row in raw_rows:
        job_id = _short_text(row.get("job_id"))
        location_id = _text(row.get("location_id"))
        sitecode = _text(row.get("sitecode")).upper()
        start_dt = _parse_datetime(row.get("outage_start_time"), date_order=date_order)
        restore_dt = _parse_datetime(row.get("power_restore_time"), date_order=date_order)
        alarm_type = _short_text(row.get("alarm_type"))
        main_cause = _short_text(row.get("main_cause"))
        subcause1 = _short_text(row.get("subcause1"))
        subcause2 = _short_text(row.get("subcause2"))
        cause_category = _cause_category(main_cause, subcause1, subcause2)
        source_duration = _parse_float(row.get("source_duration_minutes"))

        mapped = mapping.get(_norm_key(location_id))
        peanos = mapped.peanos if mapped else ()
        mapping_status = _mapping_status(location_id, peanos)
        actual = _actual_minutes(start_dt, restore_dt)
        duration_delta = _duration_delta(actual, source_duration)
        duration_consistency = _duration_consistency(actual, source_duration)
        quality = _truth_quality(
            location_id=location_id,
            peano_count=len(peanos),
            start_dt=start_dt,
            restore_dt=restore_dt,
            actual_minutes=actual,
        )
        quality_counts[quality] += 1
        mapping_counts[mapping_status] += 1
        cause_counts[cause_category] += 1
        duration_consistency_counts[duration_consistency] += 1

        peano = peanos[0] if len(peanos) == 1 else ""
        notes = _truth_notes(mapping_status, alarm_type, cause_category, main_cause, subcause1, subcause2, job_id)
        canonical = {
            "site_id": location_id,
            "peano": peano,
            "outage_start_time": _format_dt(start_dt),
            "power_restore_time": _format_dt(restore_dt),
            "actual_restoration_minutes": "" if actual is None else _format_float(actual),
            "event_number": "",
            "device_id": "",
            "feeder": "",
            "source": SOURCE_LABEL,
            "truth_source": AIS_TRUTH_SOURCE_DEFAULT,
            "truth_target": AIS_TRUTH_TARGET,
            "truth_definition": AIS_TRUTH_DEFINITION,
            "truth_quality": quality,
            "truth_notes": notes,
            "source_file": str(source_path),
            "source_row_number": str(source_row_number),
        }
        canonical_rows.append(canonical)
        if quality == "REVIEW_SHORT":
            review_rows.append(canonical)
        elif quality != "OK":
            reject_rows.append(canonical)

        audit_rows.append(
            {
                "source_row_number": str(source_row_number),
                "job_id": job_id,
                "location_id": location_id,
                "sitecode": sitecode,
                "outage_start_time": _format_dt(start_dt),
                "power_restore_time": _format_dt(restore_dt),
                "actual_restoration_minutes": "" if actual is None else _format_float(actual),
                "source_duration_minutes": "" if source_duration is None else _format_float(source_duration),
                "duration_delta_minutes": "" if duration_delta is None else _format_float(duration_delta),
                "duration_consistency": duration_consistency,
                "mapping_status": mapping_status,
                "mapped_peano_count": str(len(peanos)),
                "truth_quality": quality,
                "main_cause": main_cause,
                "subcause1": subcause1,
                "subcause2": subcause2,
                "cause_category": cause_category,
                "alarm_type": alarm_type,
            }
        )

    _write_csv(output_csv, AIS_TRUTH_COLUMNS, canonical_rows)
    _write_csv(review_csv, AIS_TRUTH_COLUMNS, review_rows)
    _write_csv(rejects_csv, AIS_TRUTH_COLUMNS, reject_rows)
    _write_csv(audit_csv, AUDIT_COLUMNS, audit_rows)

    summary = {
        "source": str(source_path),
        "meter_mapping": str(mapping_path),
        "output_csv": str(output_csv),
        "review_csv": str(review_csv),
        "rejects_csv": str(rejects_csv),
        "audit_csv": str(audit_csv),
        "report_markdown": str(report_markdown) if report_markdown else None,
        "rows": len(canonical_rows),
        "ok_rows": quality_counts.get("OK", 0),
        "review_rows": quality_counts.get("REVIEW_SHORT", 0),
        "reject_rows": len(reject_rows),
        "quality_counts": dict(sorted(quality_counts.items())),
        "mapping_counts": dict(sorted(mapping_counts.items())),
        "cause_counts": dict(sorted(cause_counts.items())),
        "duration_consistency_counts": dict(sorted(duration_consistency_counts.items())),
        "detected_columns": detected_columns,
        "date_order": date_order,
        "truth_target": AIS_TRUTH_TARGET,
        "truth_definition": AIS_TRUTH_DEFINITION,
    }
    if report_markdown:
        report = Path(report_markdown)
        report.parent.mkdir(parents=True, exist_ok=True)
        report.write_text(_render_report(summary), encoding="utf-8-sig")
    return summary


def _find_default_mapping(root: Path) -> Path:
    candidates = sorted(root.glob("Meter_ID_NE For PEA_*LatLong_R01 1.xlsx"))
    if not candidates:
        raise FileNotFoundError("Meter mapping workbook not found: Meter_ID_NE For PEA_*LatLong_R01 1.xlsx")
    return candidates[0]


def _load_meter_mapping(path: Path) -> dict[str, MappedSite]:
    rows, _ = _read_table(path, sheet="Joined")
    peanos_by_location: dict[str, set[str]] = defaultdict(set)
    site_codes_by_location: dict[str, set[str]] = defaultdict(set)
    for _, row in rows:
        location_id = _text(_pick(row, "location_id"))
        if not location_id:
            continue
        peano = _text(_pick(row, "peano")).upper()
        site_code = _text(_pick(row, "sitecode")).upper()
        key = _norm_key(location_id)
        if peano:
            peanos_by_location[key].add(peano)
        if site_code:
            site_codes_by_location[key].add(site_code)
    return {
        location_id: MappedSite(
            location_id=location_id,
            peanos=tuple(sorted(peanos)),
            site_codes=tuple(sorted(site_codes_by_location.get(location_id, set()))),
        )
        for location_id, peanos in peanos_by_location.items()
    }


def _read_alarm_rows(source: Path, sheet: str | int | None) -> tuple[list[tuple[int, dict[str, Any]]], dict[str, str]]:
    rows, detected = _read_table(source, sheet=sheet)
    output: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in rows:
        output.append(
            (
                row_number,
                {
                    "location_id": _pick(row, "location_id"),
                    "sitecode": _pick(row, "sitecode"),
                    "outage_start_time": _pick(row, "outage_start_time"),
                    "power_restore_time": _pick(row, "power_restore_time"),
                    "alarm_type": _pick(row, "alarm_type"),
                    "main_cause": _pick(row, "main_cause"),
                    "subcause1": _pick(row, "subcause1"),
                    "subcause2": _pick(row, "subcause2"),
                    "job_id": _pick(row, "job_id"),
                    "source_duration_minutes": _pick(row, "source_duration_minutes"),
                },
            )
        )
    return output, detected


def _read_table(source: Path, sheet: str | int | None = None) -> tuple[list[tuple[int, dict[str, Any]]], dict[str, str]]:
    if source.suffix.lower() in {".csv", ".txt"}:
        rows = _read_csv(source)
        if not rows:
            return [], {}
        headers = list(rows[0][1].keys())
        detected = _canonical_header_map(headers)
        mapped_rows = []
        for row_number, row in rows:
            mapped_rows.append(
                (
                    row_number,
                    {
                        canonical: row.get(original, "") if original else ""
                        for canonical, original in detected.items()
                    },
                )
            )
        return mapped_rows, detected
    elif source.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx_canonical(source, sheet=sheet)
    else:
        raise ValueError(f"Unsupported AIS add-field source type: {source.suffix}")


def _read_csv(source: Path) -> list[tuple[int, dict[str, Any]]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp874"):
        try:
            with source.open("r", encoding=encoding, newline="") as handle:
                return [
                    (index, {key: value for key, value in row.items() if key is not None})
                    for index, row in enumerate(csv.DictReader(handle), start=2)
                ]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def _read_xlsx_canonical(source: Path, sheet: str | int | None = None) -> tuple[list[tuple[int, dict[str, Any]]], dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to import AIS add-field Excel files")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if sheet is None:
            worksheet = workbook[workbook.sheetnames[0]]
        elif isinstance(sheet, int):
            worksheet = workbook[workbook.sheetnames[sheet]]
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(iterator)]
        except StopIteration:
            return [], {}
        detected = _canonical_header_map(headers)
        selected_indexes = {
            canonical: headers.index(original)
            for canonical, original in detected.items()
            if original in headers
        }
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_index, values in enumerate(iterator, start=2):
            rows.append(
                (
                    row_index,
                    {
                        canonical: values[index] if index < len(values) else ""
                        for canonical, index in selected_indexes.items()
                    },
                )
            )
        return rows, detected
    finally:
        workbook.close()


def _canonical_header_map(headers: Iterable[str]) -> dict[str, str]:
    normalized_to_original = {_normalize_header(header): str(header) for header in headers if str(header).strip()}
    aliases = {
        "location_id": ["location_id", "locationid"],
        "sitecode": ["sitecode", "site_code", "site code", "site_id", "siteid"],
        "peano": [
            # Prefer the explicit PEA meter number over generic AIS fields such as `Meter`.
            "หมายเลขเครื่องวัด PEA",
            "หมายเลขเครื่องวัดpea",
            "หมายเลขเครื่องวัด_pea",
            "หมายเลขเครื่องวัด_p_e_a",
            "peano",
            "pea_no",
            "pea_number",
            "pea number",
            "meter_no",
            "meter_number",
            "meter",
        ],
        "outage_start_time": ["firstoccurrence", "first_occurrence", "first occurred on", "first_occurred_on", "create_date", "createdate"],
        "power_restore_time": ["cleartime", "clear_time", "cleared on", "cleared_on", "done_date", "donedate"],
        "alarm_type": ["alarmname", "alarm_name", "description", "subcause1", "subcause_1"],
        "main_cause": ["maincause", "main_cause", "main cause"],
        "subcause1": ["subcause1", "subcause_1", "subcause 1", "subcase1", "subcase_1"],
        "subcause2": ["subcause2", "subcause_2", "subcause 2", "subcase", "subcase2", "subcase_2"],
        "job_id": ["jb_id", "job_id", "jobid", "ticket_id", "ticketid"],
        "source_duration_minutes": ["down_time", "downtime", "down time", "duration_minutes", "duration min"],
    }
    output: dict[str, str] = {}
    for canonical, values in aliases.items():
        output[canonical] = ""
        for alias in (canonical, *values):
            original = normalized_to_original.get(_normalize_header(alias))
            if original:
                output[canonical] = original
                break
    return output


def _pick(row: dict[str, Any], key: str) -> Any:
    return row.get(key, "")


def _mapping_status(location_id: str, peanos: tuple[str, ...]) -> str:
    if not location_id:
        return "missing_location_id"
    if not peanos:
        return "no_mapped_peano"
    if len(peanos) > 1:
        return "ambiguous_multiple_peano"
    return "matched_single_peano"


def _truth_quality(
    *,
    location_id: str,
    peano_count: int,
    start_dt: datetime | None,
    restore_dt: datetime | None,
    actual_minutes: float | None,
) -> str:
    if not location_id:
        return "MISSING_ASSET_ID"
    if start_dt is None:
        return "MISSING_OUTAGE_START"
    if restore_dt is None:
        return "MISSING_RESTORE"
    if actual_minutes is None:
        return "MISSING_RESTORE"
    if actual_minutes < 0:
        return "INVALID_NEGATIVE"
    if actual_minutes > 1440:
        return "INVALID_LONG"
    if peano_count == 0:
        return "MISSING_PEANO_MAPPING"
    if peano_count > 1:
        return "AMBIGUOUS_PEANO_MAPPING"
    if actual_minutes <= 5:
        return "REVIEW_SHORT"
    return "OK"


def _truth_notes(
    mapping_status: str,
    alarm_type: str,
    cause_category: str,
    main_cause: str,
    subcause1: str,
    subcause2: str,
    job_id: str,
) -> str:
    parts = [f"mapping_status={mapping_status}"]
    if cause_category:
        parts.append(f"cause_category={cause_category}")
    if job_id:
        parts.append(f"job_id={job_id}")
    if main_cause:
        parts.append(f"main_cause={main_cause}")
    if subcause1:
        parts.append(f"subcause1={subcause1}")
    if subcause2:
        parts.append(f"subcause2={subcause2}")
    if alarm_type:
        parts.append(f"alarm_type={alarm_type}")
    return "; ".join(parts)


def _cause_category(main_cause: str, subcause1: str, subcause2: str) -> str:
    combined = " ".join([main_cause, subcause1, subcause2]).lower()
    sub2 = subcause2.lower()
    if "pea activity" in combined or "mea/pea activity" in combined:
        return "pea_activity"
    if "have backup" in combined and ("pea" in combined or "mea" in combined):
        return "pea_have_backup"
    if "no back" in combined and "pea" in combined:
        return "pea_no_backup"
    if "ac main" in combined and ("pea" in combined or "mea" in combined):
        return "pea_ac_main_unspecified"
    if "ac main" in combined and sub2 and not any(token in sub2 for token in ("pea", "mea")):
        return "ais_equipment_or_site_fault"
    if "ac main" in combined:
        return "ac_main_failed_uncategorized"
    return "other_or_uncategorized"


def _parse_float(value: Any) -> float | None:
    text = _text(value).replace(",", "")
    if not text or text.lower() in _EMPTY_VALUES:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _duration_delta(actual_minutes: float | None, source_duration_minutes: float | None) -> float | None:
    if actual_minutes is None or source_duration_minutes is None:
        return None
    return round(actual_minutes - source_duration_minutes, 2)


def _duration_consistency(actual_minutes: float | None, source_duration_minutes: float | None) -> str:
    if actual_minutes is None:
        return "missing_actual_duration"
    if source_duration_minutes is None:
        return "missing_source_duration"
    if abs(actual_minutes - source_duration_minutes) <= 1:
        return "matches_within_1_minute"
    return "mismatch_gt_1_minute"


def _parse_datetime(value: Any, *, date_order: str) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if 20000 <= number <= 60000:
            return datetime(1899, 12, 30) + timedelta(days=number)
        return None
    text = _text(value)
    if not text or text.lower() in _EMPTY_VALUES:
        return None
    try:
        number = float(text)
        if 20000 <= number <= 60000:
            return datetime(1899, 12, 30) + timedelta(days=number)
    except ValueError:
        pass

    if date_order not in {"mdy", "dmy"}:
        raise ValueError("date_order must be 'mdy' or 'dmy'")
    slash_formats = (
        ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y")
        if date_order == "mdy"
        else ("%d/%m/%Y %I:%M:%S %p", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y")
    )
    fallback_slash_formats = (
        ("%d/%m/%Y %I:%M:%S %p", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y")
        if date_order == "mdy"
        else ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y")
    )
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        *slash_formats,
        *fallback_slash_formats,
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.year >= 2400:
                parsed = parsed.replace(year=parsed.year - 543)
            return parsed
        except ValueError:
            pass
    try:
        parsed = datetime.fromisoformat(text)
        if parsed.year >= 2400:
            parsed = parsed.replace(year=parsed.year - 543)
        return parsed.replace(tzinfo=None)
    except ValueError:
        return None


def _actual_minutes(start_dt: datetime | None, restore_dt: datetime | None) -> float | None:
    if start_dt is None or restore_dt is None:
        return None
    return round((restore_dt - start_dt).total_seconds() / 60, 2)


def _write_csv(path: str | Path, columns: Iterable[str], rows: list[dict[str, str]]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        writer.writerows({column: row.get(column, "") for column in columns} for row in rows)


def _render_report(summary: dict[str, Any]) -> str:
    lines = [
        "# AIS Add-Field Truth Import",
        "",
        "This report converts AIS AC MAIN FAIL add-field alarms into canonical AIS truth candidates for shadow evaluation.",
        "",
        "## Summary",
        "",
        f"- Source: `{Path(summary['source']).name}`",
        f"- Meter mapping: `{Path(summary['meter_mapping']).name}`",
        f"- Rows processed: {summary['rows']}",
        f"- OK sustained candidate rows: {summary['ok_rows']}",
        f"- Review rows (`<=5 min`): {summary['review_rows']}",
        f"- Reject rows: {summary['reject_rows']}",
        f"- Date parse order: `{summary['date_order']}`",
        "",
        "## Quality Counts",
        "",
        "| Truth quality | Rows |",
        "| --- | ---: |",
    ]
    for label, count in summary["quality_counts"].items():
        lines.append(f"| {label} | {count} |")
    lines.extend(["", "## Mapping Counts", "", "| Mapping status | Rows |", "| --- | ---: |"])
    for label, count in summary["mapping_counts"].items():
        lines.append(f"| {label} | {count} |")
    lines.extend(["", "## Cause Categories", "", "| Cause category | Rows |", "| --- | ---: |"])
    for label, count in summary["cause_counts"].items():
        lines.append(f"| {label} | {count} |")
    lines.extend(["", "## Duration Consistency", "", "| Check | Rows |", "| --- | ---: |"])
    for label, count in summary["duration_consistency_counts"].items():
        lines.append(f"| {label} | {count} |")
    lines.extend(
        [
            "",
            "## Outputs",
            "",
            f"- Canonical candidate truth: `{summary['output_csv']}`",
            f"- Review-only short interruptions: `{summary['review_csv']}`",
            f"- Rejects: `{summary['rejects_csv']}`",
            f"- Join audit: `{summary['audit_csv']}`",
            "",
            "## Policy",
            "",
            "- `actual_restoration_minutes = Cleartime - Firstoccurrence`",
            "- For this AIS add-field schema, `CREATE_DATE` is treated as outage start and `DONE_DATE` as restore time.",
            "- `Down Time` is used only as a consistency check; it does not replace missing restore time.",
            "- Cause/subcause is stored as context for eligibility and future model challengers; it does not replace outage/restore truth.",
            "- `<=5 minutes` stays review-only and is not used for the default model gate.",
            "- Rows without a single mapped meter PEANO are not eligible for automatic shadow truth matching.",
            "- Production AIS notification remains blocked; this is shadow evaluation input only.",
        ]
    )
    return "\n".join(lines) + "\n"


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-.#/():]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _norm_key(value: Any) -> str:
    return _text(value).upper().replace(" ", "")


def _short_text(value: Any, limit: int = 80) -> str:
    text = _text(value)
    text = re.sub(r"\s+", " ", text)
    return text[:limit]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_dt(value: datetime | None) -> str:
    return value.isoformat(sep=" ", timespec="seconds") if value else ""


def _format_float(value: float) -> str:
    if float(value).is_integer():
        return f"{value:.1f}"
    return f"{value:g}"
