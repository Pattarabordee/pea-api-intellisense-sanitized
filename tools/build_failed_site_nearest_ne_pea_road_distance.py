from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import json
import math
from pathlib import Path
import re
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from typing import Any, Iterable

from openpyxl import load_workbook


OSRM_TABLE_URL = "https://router.project-osrm.org/table/v1/driving"
DEFAULT_WORKBOOK_GLOB = "Meter_ID_NE For PEA_*LatLong_R01 1.xlsx"
DEFAULT_FAILED_SITE_CSV = Path("runtime/ais_truth_latest_candidate.csv")
DEFAULT_OFFICES_CSV = Path("runtime/ne_pea_office_locations.csv")
DEFAULT_OUTPUT = Path("runtime/ais_failed_site_nearest_ne_pea_office_road_distance.csv")
DEFAULT_SUMMARY = Path("runtime/ais_failed_site_nearest_ne_pea_office_road_distance.md")
DEFAULT_SQLITE = Path("runtime/ais_failed_site_nearest_ne_pea_office_road_distance.sqlite")
DEFAULT_OSRM_CACHE = Path("runtime/osrm_failed_site_ne_pea_office_cache.json")


@dataclass(frozen=True)
class Site:
    site_ref: str
    base_site_id: str
    site_code: str
    company: str
    source_province: str
    lat: float
    lon: float


@dataclass(frozen=True)
class Office:
    office_id: str
    office_name: str
    office_type: str
    official_address: str
    lat: float
    lon: float
    coordinate_source: str
    coordinate_url: str
    confidence: str


def build_failed_site_nearest_ne_pea_road_distance(
    failed_site_csv: str | Path,
    office_csv: str | Path,
    workbook: str | Path | None,
    output_csv: str | Path,
    summary_output: str | Path,
    sqlite_output: str | Path,
    osrm_cache: str | Path,
    *,
    batch_size: int = 10,
    office_chunk_size: int = 90,
    refresh_osrm: bool = False,
    sleep_seconds: float = 0.2,
) -> dict[str, Any]:
    failed_path = Path(failed_site_csv)
    office_path = Path(office_csv)
    workbook_path = Path(workbook) if workbook else _find_default_workbook()
    output_path = Path(output_csv)
    summary_path = Path(summary_output)
    sqlite_path = Path(sqlite_output)
    cache_path = Path(osrm_cache)

    failed_site_ids, failed_rows = load_failed_site_ids(failed_path)
    offices = load_offices(office_path)
    sites, site_stats = load_failed_site_coordinates(workbook_path, failed_site_ids)
    if not offices:
        raise RuntimeError("No valid PEA office coordinates loaded.")
    if not sites:
        raise RuntimeError("No failed AIS site coordinates loaded.")

    rows = compute_exact_nearest_by_road(
        sites,
        offices,
        batch_size=batch_size,
        office_chunk_size=office_chunk_size,
        osrm_cache_path=cache_path,
        refresh_osrm=refresh_osrm,
        sleep_seconds=sleep_seconds,
    )
    fieldnames = list(rows[0].keys())
    metadata = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "failed_site_csv": str(failed_path),
        "office_csv": str(office_path),
        "site_workbook": str(workbook_path),
        "output_csv": str(output_path),
        "summary": str(summary_path),
        "sqlite": str(sqlite_path),
        "osrm_cache": str(cache_path),
        "failed_rows": failed_rows,
        "failed_unique_site_ids": len(failed_site_ids),
        "offices": len(offices),
        "site_stats": site_stats,
        "osrm_batch_size": batch_size,
        "office_chunk_size": office_chunk_size,
    }
    write_csv_rows(output_path, fieldnames, rows)
    write_sqlite(sqlite_path, rows, metadata)
    write_summary(summary_path, rows, metadata)
    return {
        "rows": len(rows),
        "offices": len(offices),
        "failed_rows": failed_rows,
        "failed_unique_site_ids": len(failed_site_ids),
        "site_ids_with_coordinates": site_stats["site_ids_with_coordinates"],
        "site_ids_missing_coordinates": site_stats["site_ids_missing_coordinates"],
        "output": str(output_path),
        "summary": str(summary_path),
        "sqlite": str(sqlite_path),
        "osrm_cache": str(cache_path),
    }


def load_failed_site_ids(path: Path) -> tuple[set[str], int]:
    if not path.exists():
        raise FileNotFoundError(f"Failed-site CSV not found: {path}")
    rows = _read_csv(path)
    if not rows:
        return set(), 0
    site_column = _detect_column(rows[0].keys(), ("site_id", "site_ref", "Location ID", "LocationID", "location_id"))
    if not site_column:
        raise ValueError("Failed-site CSV must contain site_id, site_ref, or Location ID")
    site_ids = {_normalize_site_key(row.get(site_column)) for row in rows}
    return {site_id for site_id in site_ids if site_id}, len(rows)


def load_offices(path: Path) -> list[Office]:
    if not path.exists():
        raise FileNotFoundError(f"PEA office CSV not found: {path}")
    offices = []
    for row in _read_csv(path):
        lat = _optional_float(row.get("lat") or row.get("office_lat"))
        lon = _optional_float(row.get("lon") or row.get("office_lon"))
        if lat is None or lon is None or not _valid_thai_coord(lat, lon):
            continue
        office_id = _text(row.get("office_id")) or f"office_{len(offices) + 1:03d}"
        offices.append(
            Office(
                office_id=office_id,
                office_name=_text(row.get("office_name") or row.get("requested_office") or row.get("nearest_pea_office")),
                office_type=_text(row.get("office_type") or row.get("size") or row.get("nearest_office_size")),
                official_address=_text(row.get("official_address") or row.get("office_address")),
                lat=lat,
                lon=lon,
                coordinate_source=_text(row.get("coordinate_source") or row.get("office_coord_source")),
                coordinate_url=_text(row.get("coordinate_url") or row.get("office_coord_url")),
                confidence=_text(row.get("confidence") or row.get("coordinate_status")),
            )
        )
    return offices


def load_failed_site_coordinates(workbook_path: Path, failed_site_ids: set[str]) -> tuple[list[Site], dict[str, Any]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"AIS site coordinate workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["Joined"]
        rows = ws.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows)]
        index = {name: idx for idx, name in enumerate(headers)}
        required = ("Location ID", "Province", "Com", "LAT", "LONG")
        missing = [name for name in required if name not in index]
        if missing:
            raise ValueError(f"Missing workbook columns: {', '.join(missing)}")

        sites = []
        seen_coordinates: set[tuple[str, float, float]] = set()
        ref_counts: dict[str, int] = {}
        workbook_rows = invalid_coordinates = duplicate_site_coordinate = 0
        for row in rows:
            workbook_rows += 1
            base_site_id = _normalize_site_key(row[index["Location ID"]])
            if base_site_id not in failed_site_ids:
                continue
            lat = _optional_float(row[index["LAT"]])
            lon = _optional_float(row[index["LONG"]])
            if lat is None or lon is None or not _valid_thai_coord(lat, lon):
                invalid_coordinates += 1
                continue
            coord_key = (base_site_id, round(lat, 7), round(lon, 7))
            if coord_key in seen_coordinates:
                duplicate_site_coordinate += 1
                continue
            seen_coordinates.add(coord_key)
            ref_counts[base_site_id] = ref_counts.get(base_site_id, 0) + 1
            site_ref = base_site_id if ref_counts[base_site_id] == 1 else f"{base_site_id}__coord{ref_counts[base_site_id]}"
            sites.append(
                Site(
                    site_ref=site_ref,
                    base_site_id=base_site_id,
                    site_code=_text(row[index["SITE Code"]]) if "SITE Code" in index else "",
                    company=_text(row[index["Com"]]),
                    source_province=_text(row[index["Province"]]),
                    lat=lat,
                    lon=lon,
                )
            )
    finally:
        wb.close()

    ids_with_coordinates = {site.base_site_id for site in sites}
    return sites, {
        "workbook_rows": workbook_rows,
        "failed_site_coordinate_rows": len(sites),
        "site_ids_with_coordinates": len(ids_with_coordinates),
        "site_ids_missing_coordinates": len(failed_site_ids - ids_with_coordinates),
        "invalid_coordinates": invalid_coordinates,
        "duplicate_site_coordinate": duplicate_site_coordinate,
    }


def compute_exact_nearest_by_road(
    sites: list[Site],
    offices: list[Office],
    batch_size: int,
    office_chunk_size: int,
    osrm_cache_path: Path,
    refresh_osrm: bool = False,
    sleep_seconds: float = 0.2,
) -> list[dict[str, Any]]:
    signature = _office_signature(offices)
    cache = _load_json(osrm_cache_path, {"office_signature": signature, "items": {}})
    if not isinstance(cache, dict) or cache.get("office_signature") != signature or refresh_osrm:
        cache = {"office_signature": signature, "items": {}}
    items = cache.setdefault("items", {})
    if not isinstance(items, dict):
        items = {}
        cache["items"] = items

    results_by_key: dict[str, dict[str, Any]] = {}
    missing_sites = []
    for site in sites:
        key = _site_cache_key(site)
        if key in items:
            results_by_key[key] = dict(items[key])
        else:
            missing_sites.append(site)

    site_batches = list(_chunks(missing_sites, batch_size))
    office_chunks = list(_chunks(offices, office_chunk_size))
    total_requests = len(site_batches) * len(office_chunks)
    request_count = 0
    for site_batch_number, site_batch in enumerate(site_batches, start=1):
        best_by_site: dict[str, dict[str, Any] | None] = {_site_cache_key(site): None for site in site_batch}
        for office_chunk in office_chunks:
            request_count += 1
            distances, durations = osrm_table_batch(site_batch, office_chunk)
            for site_idx, site in enumerate(site_batch):
                candidate = _best_from_matrix_row(site, office_chunk, distances[site_idx], durations[site_idx] if durations else [])
                key = _site_cache_key(site)
                best_by_site[key] = _choose_better(best_by_site[key], candidate)
            if sleep_seconds:
                time.sleep(sleep_seconds)
            print(
                f"OSRM request {request_count}/{total_requests}: site batch {site_batch_number}/{len(site_batches)}",
                flush=True,
            )
        for site in site_batch:
            key = _site_cache_key(site)
            best = best_by_site[key] or _straight_line_fallback(site, offices)
            items[key] = best
            results_by_key[key] = best
        _write_json(osrm_cache_path, cache)

    rows = []
    office_by_id = {office.office_id: office for office in offices}
    for site in sites:
        nearest = results_by_key[_site_cache_key(site)]
        office = office_by_id[str(nearest["nearest_office_id"])]
        rows.append(
            {
                "site_ref": site.site_ref,
                "site_code": site.site_code,
                "company": site.company,
                "source_province": site.source_province,
                "site_lat": _fmt(site.lat, 7),
                "site_lon": _fmt(site.lon, 7),
                "nearest_pea_office": office.office_name,
                "nearest_office_size": office.office_type,
                "official_office_name": office.office_name,
                "office_address": office.official_address,
                "office_lat": _fmt(office.lat, 7),
                "office_lon": _fmt(office.lon, 7),
                "road_distance_km": _fmt(nearest.get("road_distance_km")),
                "straight_line_km": _fmt(nearest.get("straight_line_km")),
                "osrm_duration_min": _fmt(nearest.get("osrm_duration_min"), 1),
                "route_status": nearest.get("route_status", ""),
                "office_coord_source": office.coordinate_source,
                "office_coord_url": office.coordinate_url,
                "office_confidence": office.confidence,
            }
        )
    return rows


def osrm_table_batch(
    sites: list[Site],
    offices: list[Office],
    retries: int = 4,
    timeout: int = 90,
) -> tuple[list[list[float | None]], list[list[float | None]]]:
    coords = [f"{site.lon:.7f},{site.lat:.7f}" for site in sites]
    coords.extend(f"{office.lon:.7f},{office.lat:.7f}" for office in offices)
    source_indexes = ";".join(str(idx) for idx in range(len(sites)))
    destination_indexes = ";".join(str(len(sites) + idx) for idx in range(len(offices)))
    url = f"{OSRM_TABLE_URL}/{';'.join(coords)}?sources={source_indexes}&destinations={destination_indexes}&annotations=distance,duration"

    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            text = _fetch_text(url, timeout=timeout)
            payload = json.loads(text)
            if payload.get("code") != "Ok":
                raise RuntimeError(f"OSRM returned {payload.get('code')}: {payload.get('message')}")
            return payload["distances"], payload.get("durations", [])
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, RuntimeError) as exc:
            last_error = exc
            if attempt == retries:
                break
            time.sleep(min(20, 2**attempt))
    raise RuntimeError(f"OSRM table request failed after {retries} attempts: {last_error}")


def _best_from_matrix_row(
    site: Site,
    offices: list[Office],
    distances_m: list[float | None],
    durations_s: list[float | None],
) -> dict[str, Any] | None:
    best_idx = None
    best_distance = None
    for idx, distance in enumerate(distances_m):
        if distance is None:
            continue
        if best_distance is None or distance < best_distance:
            best_idx = idx
            best_distance = distance
    if best_idx is None or best_distance is None:
        return None
    office = offices[best_idx]
    duration = durations_s[best_idx] if durations_s and durations_s[best_idx] is not None else None
    return {
        "nearest_office_id": office.office_id,
        "road_distance_km": round(best_distance / 1000, 3),
        "straight_line_km": round(haversine_km(site.lat, site.lon, office.lat, office.lon), 3),
        "osrm_duration_min": round(duration / 60, 1) if duration is not None else None,
        "route_status": "ok",
    }


def _choose_better(current: dict[str, Any] | None, candidate: dict[str, Any] | None) -> dict[str, Any] | None:
    if candidate is None:
        return current
    if current is None:
        return candidate
    current_distance = _optional_float(current.get("road_distance_km"))
    candidate_distance = _optional_float(candidate.get("road_distance_km"))
    if current_distance is None:
        return candidate
    if candidate_distance is None:
        return current
    return candidate if candidate_distance < current_distance else current


def _straight_line_fallback(site: Site, offices: list[Office]) -> dict[str, Any]:
    best_office = min(offices, key=lambda office: haversine_km(site.lat, site.lon, office.lat, office.lon))
    return {
        "nearest_office_id": best_office.office_id,
        "road_distance_km": None,
        "straight_line_km": round(haversine_km(site.lat, site.lon, best_office.lat, best_office.lon), 3),
        "osrm_duration_min": None,
        "route_status": "osrm_no_route",
    }


def write_csv_rows(path: str | Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_sqlite(path: str | Path, rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(output)
    try:
        conn.execute("DROP TABLE IF EXISTS nearest_failed_site_office")
        conn.execute("DROP TABLE IF EXISTS metadata")
        conn.execute(
            """
            CREATE TABLE nearest_failed_site_office (
                site_ref TEXT,
                site_code TEXT,
                company TEXT,
                source_province TEXT,
                site_lat REAL,
                site_lon REAL,
                nearest_pea_office TEXT,
                nearest_office_size TEXT,
                official_office_name TEXT,
                office_address TEXT,
                office_lat REAL,
                office_lon REAL,
                road_distance_km REAL,
                straight_line_km REAL,
                osrm_duration_min REAL,
                route_status TEXT,
                office_coord_source TEXT,
                office_coord_url TEXT,
                office_confidence TEXT
            )
            """
        )
        conn.execute("CREATE TABLE metadata (key TEXT PRIMARY KEY, value TEXT)")
        if rows:
            field_count = len(rows[0])
            conn.executemany(
                f"INSERT INTO nearest_failed_site_office VALUES ({','.join(['?'] * field_count)})",
                ([row.get(column, "") for column in rows[0].keys()] for row in rows),
            )
        conn.executemany(
            "INSERT INTO metadata VALUES (?,?)",
            [(key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else str(value)) for key, value in metadata.items()],
        )
        conn.commit()
    finally:
        conn.close()


def write_summary(path: str | Path, rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    road_values = [value for value in (_optional_float(row.get("road_distance_km")) for row in rows) if value is not None]
    route_status_counts: dict[str, int] = {}
    office_counts: dict[str, int] = {}
    for row in rows:
        route_status_counts[str(row.get("route_status") or "")] = route_status_counts.get(str(row.get("route_status") or ""), 0) + 1
        office = str(row.get("nearest_pea_office") or "")
        office_counts[office] = office_counts.get(office, 0) + 1
    top_offices = sorted(office_counts.items(), key=lambda item: (-item[1], item[0]))[:12]
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Failed AIS Site to NE PEA Office Road Distance",
        "",
        "Scope: failed AIS sites from the source CSV routed to the nearest office among the supplied NE PEA office CSV.",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "| --- | ---: |",
        f"| Failed source rows | {metadata['failed_rows']} |",
        f"| Failed unique site IDs | {metadata['failed_unique_site_ids']} |",
        f"| Site IDs with coordinates | {metadata['site_stats']['site_ids_with_coordinates']} |",
        f"| Site IDs missing coordinates | {metadata['site_stats']['site_ids_missing_coordinates']} |",
        f"| Routed coordinate rows | {len(rows)} |",
        f"| Offices considered | {metadata['offices']} |",
        f"| Max one-way road distance km | {_blank(max(road_values) if road_values else None)} |",
        "",
        "## Route Status",
        "",
        "| Status | Rows |",
        "| --- | ---: |",
    ]
    for status, count in sorted(route_status_counts.items()):
        lines.append(f"| `{status or '<blank>'}` | {count} |")
    lines.extend(["", "## Top Nearest Offices", "", "| Office | Site coordinate rows |", "| --- | ---: |"])
    for office, count in top_offices:
        lines.append(f"| {office} | {count} |")
    lines.extend(
        [
            "",
            "## Guardrails",
            "",
            "- Output excludes PEANO, meter numbers, customer identity, raw WebEx text, room IDs, tokens, and secrets.",
            "- This is a feature lookup only; AIS outage/restore remains the truth label.",
            "- Rows without AIS site coordinates cannot be routed and are counted as missing coordinates.",
        ]
    )
    output.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius_km = 6371.0088
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)
    a = (
        math.sin(d_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    )
    return 2 * radius_km * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _fetch_text(url: str, timeout: int = 60) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": "pea-ais-etr-distance-feature/1.0"})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read().decode("utf-8")


def _office_signature(offices: list[Office]) -> str:
    material = "|".join(f"{office.office_id}:{office.lat:.7f},{office.lon:.7f}" for office in offices)
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def _site_cache_key(site: Site) -> str:
    return f"{site.site_ref}|{site.lat:.7f},{site.lon:.7f}"


def _read_csv(path: Path) -> list[dict[str, str]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp874"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [{key: (value or "") for key, value in row.items() if key is not None} for row in csv.DictReader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def _chunks(items: list[Any], size: int) -> Iterable[list[Any]]:
    for start in range(0, len(items), size):
        yield items[start : start + size]


def _detect_column(columns: Iterable[str], candidates: Iterable[str]) -> str:
    normalized = {_normalize_header(column): column for column in columns}
    for candidate in candidates:
        column = normalized.get(_normalize_header(candidate))
        if column:
            return column
    return ""


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-.#/()]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _normalize_site_key(value: Any) -> str:
    text = _text(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", "", text).upper()


def _optional_float(value: Any) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _valid_thai_coord(lat: float, lon: float) -> bool:
    return 5 <= lat <= 21 and 97 <= lon <= 106


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _fmt(value: Any, digits: int = 3) -> str:
    numeric = _optional_float(value)
    if numeric is None:
        return ""
    rounded = round(numeric, digits)
    return str(int(rounded)) if rounded == int(rounded) else str(rounded).rstrip("0").rstrip(".")


def _blank(value: Any) -> str:
    if value is None or value == "":
        return ""
    return _fmt(value) if isinstance(value, (float, int)) else str(value)


def _find_default_workbook() -> Path:
    matches = list(Path(".").glob(DEFAULT_WORKBOOK_GLOB))
    if not matches:
        raise FileNotFoundError(f"Cannot find workbook matching {DEFAULT_WORKBOOK_GLOB}")
    return matches[0]


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Route failed AIS sites to the nearest NE PEA office by OSRM road distance."
    )
    parser.add_argument("--failed-site-csv", default=str(DEFAULT_FAILED_SITE_CSV))
    parser.add_argument("--office-csv", default=str(DEFAULT_OFFICES_CSV))
    parser.add_argument("--workbook", default=None)
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--summary", default=str(DEFAULT_SUMMARY))
    parser.add_argument("--sqlite", default=str(DEFAULT_SQLITE))
    parser.add_argument("--osrm-cache", default=str(DEFAULT_OSRM_CACHE))
    parser.add_argument("--batch-size", type=int, default=10)
    parser.add_argument("--office-chunk-size", type=int, default=90)
    parser.add_argument("--sleep-seconds", type=float, default=0.2)
    parser.add_argument("--refresh-osrm", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    if args.batch_size < 1:
        parser.error("--batch-size must be >= 1")
    if args.office_chunk_size < 1:
        parser.error("--office-chunk-size must be >= 1")
    if args.batch_size + args.office_chunk_size > 100:
        parser.error("--batch-size plus --office-chunk-size must be <= 100 for public OSRM safety")
    result = build_failed_site_nearest_ne_pea_road_distance(
        args.failed_site_csv,
        args.office_csv,
        args.workbook,
        args.output,
        args.summary,
        args.sqlite,
        args.osrm_cache,
        batch_size=args.batch_size,
        office_chunk_size=args.office_chunk_size,
        refresh_osrm=args.refresh_osrm,
        sleep_seconds=args.sleep_seconds,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
