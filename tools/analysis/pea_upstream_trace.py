"""
PEA GIS — Upstream Trace (ฉบับแก้ไขจาก Schema จริง)
======================================================
รัน Script นี้บนเครื่องที่เชื่อมต่ออินเทอร์เน็ต (หรือ Intranet PEA)

สิ่งที่ script นี้ทำ:
  สำหรับ Meter แต่ละลูกใน Excel → ค้นหา upstream ทั้งหมด:
    1. หม้อแปลงไฟฟ้า (Layer 17) ที่จ่ายไฟให้ Meter นั้น
    2. Recloser (Layer 14) ที่คุ้มครอง Feeder
    3. Switch/Fuse แรงสูง (Layer 16) ในเส้นทางเดียวกัน
    4. Circuit Breaker (Layer 11) ต้นทางของ Feeder

แก้ไขจาก Schema จริง:
  - FEEDERID ใน Layer 26 มักว่าง → ใช้ Geometry-based join แทน
  - L29 ไม่มี Geometry → bridge ผ่าน L26 ด้วย PEANO
  - kVA ของ Transformer = RATEKVA (ไม่ใช่ RATEDKVA)
  - L16 มี SUBTYPECODE ดึง domain มา map ประเภทอุปกรณ์อัตโนมัติ
  - L11 มี HVTRANSFORMER_ID → shortcut ชี้ตรงไปหา HV Transformer

Dependencies:
    pip install openpyxl

Usage:
    python pea_upstream_trace.py

Input:  เลขม_เตอร__PEA_สกลนคร.xlsx  (วางไว้ใน folder เดียวกัน)
Output: upstream_result.xlsx
"""

import json
import math
import time
import urllib.parse
import urllib.request
from datetime import datetime

# ─────────────────────────────────────────────────────
#  CONFIG — แก้ค่าเหล่านี้ถ้าจำเป็น
# ─────────────────────────────────────────────────────
INPUT_FILE  = "PEA_no_Sakon.xlsx"
OUTPUT_FILE = "upstream_result.xlsx"
BASE_URL    = "https://gisne1.pea.co.th/arcgis/rest/services"
# BASE_URL  = "http://172.16.184.233/arcgis/rest/services"  # Intranet PEA

REQUEST_DELAY   = 0.35   # วินาที ระหว่าง request (ไม่ให้ถี่เกินไป)
REQUEST_TIMEOUT = 45     # วินาที timeout ต่อ request
REQUEST_RETRIES = 3      # จำนวนครั้ง retry

# Layer IDs
L_LV_METER      = ("PEA",       26)   # DS_LowVoltageMeter
L_METER_QUERY   = ("PEA_QUERY", 29)   # DS_Meter + ชื่อลูกค้า
L_TRANSFORMER   = ("PEA",       17)   # DS_Transformer
L_RECLOSER      = ("PEA",       14)   # DS_Recloser
L_SWITCH        = ("PEA",       16)   # DS_Switch / Fuse
L_CB            = ("PEA",       11)   # DS_CircuitBreaker

# ─────────────────────────────────────────────────────
#  HTTP HELPER
# ─────────────────────────────────────────────────────
def fetch(url: str) -> dict:
    """GET request พร้อม retry และ exponential back-off"""
    for attempt in range(REQUEST_RETRIES):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "PEA-UpstreamTrace/2.0"}
            )
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as r:
                return json.loads(r.read().decode("utf-8", errors="replace"))
        except Exception as e:
            wait = 2 ** attempt
            print(f"    ⚠ attempt {attempt+1}/{REQUEST_RETRIES} failed: {e}"
                  f" — retry in {wait}s")
            if attempt < REQUEST_RETRIES - 1:
                time.sleep(wait)
    return {}


def query(service: str, layer_id: int, params: dict) -> list:
    """Query Feature Layer — คืน list of attribute dicts"""
    base = f"{BASE_URL}/{service}/MapServer/{layer_id}/query"
    params.setdefault("f", "pjson")
    url = base + "?" + urllib.parse.urlencode(params)
    data = fetch(url)
    time.sleep(REQUEST_DELAY)
    return data.get("features", [])


def query_attrs(service, layer_id, where, fields="*", geometry=False) -> list:
    """Query แบบง่าย — คืน list of attribute dicts"""
    feats = query(service, layer_id, {
        "where":          where,
        "outFields":      fields,
        "returnGeometry": "true" if geometry else "false",
    })
    return [f.get("attributes", {}) for f in feats]


def query_one_with_geom(service, layer_id, where, fields="*") -> tuple:
    """Query แถวแรกพร้อม geometry — คืน (attrs, x, y) หรือ (None, None, None)"""
    feats = query(service, layer_id, {
        "where":          where,
        "outFields":      fields,
        "returnGeometry": "true",
        "resultRecordCount": 1,
    })
    if not feats:
        return None, None, None
    f    = feats[0]
    geom = f.get("geometry") or {}
    return f.get("attributes", {}), geom.get("x"), geom.get("y")


def count_records(service, layer_id, where="1=1") -> int:
    """นับจำนวน records"""
    data = fetch(
        f"{BASE_URL}/{service}/MapServer/{layer_id}/query?"
        + urllib.parse.urlencode({
            "where": where, "returnCountOnly": "true", "f": "pjson"
        })
    )
    time.sleep(REQUEST_DELAY)
    return data.get("count", -1)

# ─────────────────────────────────────────────────────
#  STEP 0 — ดึง Domain ของ Layer 16 SUBTYPECODE
#           เพื่อแยกประเภทอุปกรณ์ใน DS_Switch
# ─────────────────────────────────────────────────────
def load_switch_subtypes() -> dict:
    """
    ดึง Subtype domain จาก Layer 16
    คืน dict เช่น {1: 'Switch', 2: 'Fuse Cutout', 3: 'Sectionalizer', ...}
    """
    print("  ดึง Switch subtypes ...")
    url = f"{BASE_URL}/PEA/MapServer/16?f=pjson"
    info = fetch(url)
    time.sleep(REQUEST_DELAY)

    subtypes = {}
    for st in info.get("subtypes", []):
        code = st.get("code")
        name = st.get("name", f"subtype_{code}")
        if code is not None:
            subtypes[code] = name

    if not subtypes:
        # fallback — ถ้า API ไม่คืน subtypes ใช้ค่า default
        print("    ⚠ ไม่พบ subtype domain — ใช้ค่า default")
        subtypes = {
            1: "Switch",
            2: "Fuse Cutout",
            3: "Sectionalizer",
            4: "RMU",
            5: "ATS",
        }
    else:
        print(f"    พบ {len(subtypes)} subtype: {subtypes}")
    return subtypes

# ─────────────────────────────────────────────────────
#  STEP 1 — หา Meter ใน Layer 26 จาก PEANO
#           คืน (attrs, x, y)
# ─────────────────────────────────────────────────────
def find_meter_geom(peano: str) -> tuple:
    """
    ค้นหา Meter ใน Layer 26 ด้วย PEANO
    คืน (attrs_dict, x_web_mercator, y_web_mercator)
    """
    where = f"PEANO='{peano}'"
    attrs, x, y = query_one_with_geom(
        *L_LV_METER, where,
        fields="PEANO,ACCOUNTNUMBER,FEEDERID,INSTALLATIONID,LOCATION"
    )
    return attrs, x, y

# ─────────────────────────────────────────────────────
#  STEP 2 — หา Transformer จาก Geometry (nearest)
#
#  ใช้ FEEDERID ถ้ามี มิเช่นนั้นใช้ proximity ล้วน
#  คืน attrs ของ Transformer ที่ใกล้ที่สุด
# ─────────────────────────────────────────────────────
def dist2(ax, ay, bx, by) -> float:
    return (ax - bx) ** 2 + (ay - by) ** 2


def find_nearest_transformer(mx: float, my: float, feeder_id: str = None,
                             radius_m: float = 2000) -> dict:
    """
    หา Transformer ที่ใกล้ Meter (mx, my) ที่สุด
    ถ้ามี feeder_id: จำกัดค้นใน Feeder นั้น
    ถ้าไม่มี: ค้นในรัศมี radius_m เมตร (Web Mercator unit ≈ เมตร)
    คืน attribute dict ของ Transformer ที่ใกล้ที่สุด หรือ {}
    """
    if feeder_id:
        where = f"FEEDERID='{feeder_id}'"
    else:
        # ค้นด้วย envelope (bounding box) รอบ Meter
        xmin, xmax = mx - radius_m, mx + radius_m
        ymin, ymax = my - radius_m, my + radius_m
        where = "1=1"   # filter ด้วย geometry ด้านล่าง

    params = {
        "where":          where,
        "outFields":      "FACILITYID,FEEDERID,RATEKVA,EXISTINGKVA,"
                          "LOCATION,PHASEDESIGNATION,SUBTYPECODE,"
                          "NUMBEROFUSER,PEANO,GLOBALID",
        "returnGeometry": "true",
        "f":              "pjson",
    }

    # ถ้าไม่มี feeder ให้ใช้ spatial filter envelope
    if not feeder_id:
        params["geometry"] = json.dumps({
            "xmin": xmin, "ymin": ymin,
            "xmax": xmax, "ymax": ymax,
            "spatialReference": {"wkid": 102100}
        })
        params["geometryType"]    = "esriGeometryEnvelope"
        params["spatialRel"]      = "esriSpatialRelIntersects"
        params["inSR"]            = "102100"

    feats = query(*L_TRANSFORMER, params)
    time.sleep(REQUEST_DELAY)

    if not feats:
        return {}

    nearest, best_d = None, float("inf")
    for f in feats:
        g = f.get("geometry") or {}
        if g.get("x") is None:
            continue
        d = dist2(mx, my, g["x"], g["y"])
        if d < best_d:
            best_d, nearest = d, f.get("attributes", {})

    return nearest or {}

# ─────────────────────────────────────────────────────
#  STEP 3 — หา Recloser ใน Feeder เดียวกัน
# ─────────────────────────────────────────────────────
def find_reclosers(feeder_id: str) -> list:
    """คืน list of attrs dicts ของ Recloser ทุกตัวใน Feeder"""
    rows = query_attrs(
        *L_RECLOSER,
        where=f"FEEDERID='{feeder_id}'",
        fields="FACILITYID,FEEDERID,OPERATIONTYPE,LOCATION,"
               "MAXCONTINUOUSCURRENT,NORMALPOSITION,NUMBEROFUSER"
    )
    time.sleep(REQUEST_DELAY)
    return rows

# ─────────────────────────────────────────────────────
#  STEP 4 — หา Switch/Fuse แรงสูงใน Feeder เดียวกัน
# ─────────────────────────────────────────────────────
def find_switches(feeder_id: str, subtype_map: dict) -> list:
    """คืน list of attrs dicts ของ Switch/Fuse ใน Feeder"""
    rows = query_attrs(
        *L_SWITCH,
        where=f"FEEDERID='{feeder_id}'",
        fields="FACILITYID,FEEDERID,SUBTYPECODE,OPERATIONTYPE,"
               "LOCATION,MAXCONTINUOUSCURRENT,NORMALSTATUS,NUMBEROFUSER"
    )
    time.sleep(REQUEST_DELAY)
    # เพิ่ม SUBTYPENAME สำหรับความเข้าใจ
    for r in rows:
        code = r.get("SUBTYPECODE")
        r["SUBTYPENAME"] = subtype_map.get(code, f"code_{code}")
    return rows

# ─────────────────────────────────────────────────────
#  STEP 5 — หา Circuit Breaker ต้นทางของ Feeder
# ─────────────────────────────────────────────────────
def find_circuit_breakers(feeder_id: str) -> list:
    """คืน list of attrs dicts ของ CB ที่ตรงกับ FEEDERID"""
    rows = query_attrs(
        *L_CB,
        where=f"FEEDERID='{feeder_id}'",
        fields="FACILITYID,FEEDERID,OPERATIONTYPE,LOCATION,"
               "MAXCONTINUOUSCURRENT,HVTRANSFORMER_ID,NORMALSTATUS,NUMBEROFUSER"
    )
    time.sleep(REQUEST_DELAY)
    return rows

# ─────────────────────────────────────────────────────
#  STEP 6 — ดึงชื่อลูกค้าจาก L29 ด้วย PEANO
# ─────────────────────────────────────────────────────
def find_customer_info(peano: str) -> dict:
    """คืน attrs dict จาก PEA_QUERY Layer 29"""
    rows = query_attrs(
        *L_METER_QUERY,
        where=f"PEANO='{peano}'",
        fields="PEANO,CUSTOMERNAME,ACCOUNTNUMBER,ADDRESS,MOO,"
               "TUMBOL,AMPHOE,CHANGWAT,METERTYPE,USERTYPE,VOLTAGE"
    )
    return rows[0] if rows else {}

# ─────────────────────────────────────────────────────
#  MAIN TRACE — รวม step ทั้งหมด สำหรับ 1 Meter
# ─────────────────────────────────────────────────────
def trace_upstream(peano: str, subtype_map: dict) -> dict:
    """
    ทำ Upstream Trace สำหรับ 1 Meter (PEANO)
    คืน dict ผลลัพธ์พร้อมรายละเอียดอุปกรณ์ upstream ทั้งหมด
    """
    result = {
        # ── ข้อมูล Meter ──
        "PEANO":             peano,
        "METER_LOCATION":    None,
        "METER_FEEDERID":    None,
        "METER_X":           None,
        "METER_Y":           None,
        # ── ชื่อลูกค้า ──
        "CUSTOMER_NAME":     None,
        "CUSTOMER_TYPE":     None,
        "METER_TYPE":        None,
        "ADDRESS_MOO":       None,
        "ADDRESS_TUMBOL":    None,
        "ADDRESS_AMPHOE":    None,
        # ── Transformer ──
        "TX_FACILITYID":     None,
        "TX_PEANO":          None,
        "TX_FEEDERID":       None,
        "TX_RATEKVA":        None,
        "TX_EXISTINGKVA":    None,
        "TX_NUMUSER":        None,
        "TX_LOCATION":       None,
        # ── Recloser (อาจมีหลายตัว — เก็บเป็น string) ──
        "RC_COUNT":          0,
        "RC_FACILITYIDS":    None,
        "RC_LOCATIONS":      None,
        "RC_OPTYPES":        None,
        # ── Switch/Fuse ──
        "SW_COUNT":          0,
        "SW_FACILITYIDS":    None,
        "SW_SUBTYPENAMES":   None,
        "SW_LOCATIONS":      None,
        # ── Circuit Breaker ──
        "CB_COUNT":          0,
        "CB_FACILITYIDS":    None,
        "CB_LOCATIONS":      None,
        "CB_HVTX_IDS":       None,
        # ── สถานะ ──
        "TRACE_STATUS":      "OK",
        "TRACE_NOTE":        None,
    }

    notes = []

    # ── Step 1: หา Meter geometry ──────────────────────
    meter_attrs, mx, my = find_meter_geom(peano)

    if mx is None or my is None:
        result["TRACE_STATUS"] = "NO_METER"
        result["TRACE_NOTE"]   = "ไม่พบ Meter ใน Layer 26"
        return result

    result["METER_X"]        = round(mx, 2)
    result["METER_Y"]        = round(my, 2)
    result["METER_LOCATION"] = meter_attrs.get("LOCATION")
    feeder_id                = meter_attrs.get("FEEDERID")

    if feeder_id:
        result["METER_FEEDERID"] = feeder_id
    else:
        notes.append("FEEDERID ว่างใน L26 — ใช้ Geometry join")

    # ── Step 2: ดึงชื่อลูกค้าจาก L29 ────────────────────
    cust = find_customer_info(peano)
    result["CUSTOMER_NAME"]  = cust.get("CUSTOMERNAME")
    result["CUSTOMER_TYPE"]  = cust.get("USERTYPE")
    result["METER_TYPE"]     = cust.get("METERTYPE")
    result["ADDRESS_MOO"]    = cust.get("MOO")
    result["ADDRESS_TUMBOL"] = cust.get("TUMBOL")
    result["ADDRESS_AMPHOE"] = cust.get("AMPHOE")

    # ── Step 3: หา Transformer ──────────────────────────
    tx = find_nearest_transformer(mx, my, feeder_id)

    if tx:
        result["TX_FACILITYID"]  = tx.get("FACILITYID")
        result["TX_PEANO"]       = tx.get("PEANO")
        result["TX_FEEDERID"]    = tx.get("FEEDERID")
        result["TX_RATEKVA"]     = tx.get("RATEKVA")
        result["TX_EXISTINGKVA"] = tx.get("EXISTINGKVA")
        result["TX_NUMUSER"]     = tx.get("NUMBEROFUSER")
        result["TX_LOCATION"]    = tx.get("LOCATION")
        # ถ้ายังไม่มี feeder ให้ดึงจาก Transformer
        if not feeder_id and tx.get("FEEDERID"):
            feeder_id = tx["FEEDERID"]
            result["METER_FEEDERID"] = feeder_id
            notes.append(f"ได้ FEEDERID={feeder_id} จาก Transformer")
    else:
        notes.append("ไม่พบ Transformer ในบริเวณใกล้เคียง")
        result["TRACE_STATUS"] = "NO_TX"

    # ── Step 4: หา Recloser, Switch, CB (ต้องมี FEEDERID) ──
    if feeder_id:
        # Recloser
        rcs = find_reclosers(feeder_id)
        result["RC_COUNT"]       = len(rcs)
        result["RC_FACILITYIDS"] = " | ".join(
            str(r.get("FACILITYID","")) for r in rcs if r.get("FACILITYID"))
        result["RC_LOCATIONS"]   = " | ".join(
            str(r.get("LOCATION","")) for r in rcs if r.get("LOCATION"))
        result["RC_OPTYPES"]     = " | ".join(
            str(r.get("OPERATIONTYPE","")) for r in rcs if r.get("OPERATIONTYPE"))

        # Switch / Fuse
        sws = find_switches(feeder_id, subtype_map)
        result["SW_COUNT"]       = len(sws)
        result["SW_FACILITYIDS"] = " | ".join(
            str(s.get("FACILITYID","")) for s in sws if s.get("FACILITYID"))
        result["SW_SUBTYPENAMES"]= " | ".join(
            str(s.get("SUBTYPENAME","")) for s in sws if s.get("SUBTYPENAME"))
        result["SW_LOCATIONS"]   = " | ".join(
            str(s.get("LOCATION","")) for s in sws if s.get("LOCATION"))

        # Circuit Breaker
        cbs = find_circuit_breakers(feeder_id)
        result["CB_COUNT"]       = len(cbs)
        result["CB_FACILITYIDS"] = " | ".join(
            str(c.get("FACILITYID","")) for c in cbs if c.get("FACILITYID"))
        result["CB_LOCATIONS"]   = " | ".join(
            str(c.get("LOCATION","")) for c in cbs if c.get("LOCATION"))
        result["CB_HVTX_IDS"]   = " | ".join(
            str(c.get("HVTRANSFORMER_ID","")) for c in cbs
            if c.get("HVTRANSFORMER_ID"))
    else:
        result["TRACE_STATUS"] = "NO_FEEDER"
        notes.append("ไม่สามารถระบุ FEEDERID — ข้ามการหา RC/SW/CB")

    if notes:
        result["TRACE_NOTE"] = " | ".join(notes)

    return result

# ─────────────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────────────
def save_results(rows: list, path: str):
    try:
        import openpyxl
        from openpyxl.styles import (Alignment, Border, Font,
                                     PatternFill, Side)
    except ImportError:
        print("⚠  pip install openpyxl แล้วรันใหม่")
        return

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Upstream Trace"

    # ── Styles ──────────────────────────────────────────
    thin   = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP   = Alignment(wrap_text=True, vertical="top")
    BODY   = Font(name="Arial", size=9)

    def hdr_style(color_hex: str) -> dict:
        return {
            "font":      Font(name="Arial", size=9, bold=True,
                              color="FFFFFF"),
            "fill":      PatternFill("solid", fgColor=color_hex),
            "border":    BORDER,
            "alignment": Alignment(horizontal="center", vertical="center",
                                   wrap_text=True),
        }

    def apply(cell, **kwargs):
        for k, v in kwargs.items():
            setattr(cell, k, v)

    # ── Column definitions ────────────────────────────
    # (header_text, field_key, color_group, width)
    COLS = [
        # Meter info
        ("PEANO",             "PEANO",             "1F4E79", 14),
        ("สถานที่ Meter",     "METER_LOCATION",    "1F4E79", 22),
        ("Feeder ID",         "METER_FEEDERID",    "1F4E79", 10),
        # Customer
        ("ชื่อลูกค้า",        "CUSTOMER_NAME",     "375623", 24),
        ("ประเภทผู้ใช้",      "CUSTOMER_TYPE",     "375623", 12),
        ("ประเภทมิเตอร์",     "METER_TYPE",        "375623", 12),
        ("หมู่",              "ADDRESS_MOO",       "375623", 8),
        ("ตำบล",              "ADDRESS_TUMBOL",    "375623", 14),
        ("อำเภอ",             "ADDRESS_AMPHOE",    "375623", 12),
        # Transformer
        ("TX: FACILITYID",    "TX_FACILITYID",     "833C11", 14),
        ("TX: PEANO",         "TX_PEANO",          "833C11", 14),
        ("TX: Feeder",        "TX_FEEDERID",       "833C11", 10),
        ("TX: RATEKVA",       "TX_RATEKVA",        "833C11", 9),
        ("TX: Load kVA",      "TX_EXISTINGKVA",    "833C11", 9),
        ("TX: จำนวนผู้ใช้",   "TX_NUMUSER",        "833C11", 10),
        ("TX: สถานที่",       "TX_LOCATION",       "833C11", 22),
        # Recloser
        ("RC: จำนวน",         "RC_COUNT",          "7B2C2C", 7),
        ("RC: FACILITYID",    "RC_FACILITYIDS",    "7B2C2C", 22),
        ("RC: สถานที่",       "RC_LOCATIONS",      "7B2C2C", 22),
        ("RC: ประเภทควบคุม",  "RC_OPTYPES",        "7B2C2C", 12),
        # Switch / Fuse
        ("SW: จำนวน",         "SW_COUNT",          "44336B", 7),
        ("SW: FACILITYID",    "SW_FACILITYIDS",    "44336B", 22),
        ("SW: ประเภท",        "SW_SUBTYPENAMES",   "44336B", 22),
        ("SW: สถานที่",       "SW_LOCATIONS",      "44336B", 22),
        # Circuit Breaker
        ("CB: จำนวน",         "CB_COUNT",          "0D4E6E", 7),
        ("CB: FACILITYID",    "CB_FACILITYIDS",    "0D4E6E", 22),
        ("CB: สถานที่",       "CB_LOCATIONS",      "0D4E6E", 22),
        ("CB: HV TX ID",      "CB_HVTX_IDS",       "0D4E6E", 14),
        # Status
        ("สถานะ",             "TRACE_STATUS",      "555555", 10),
        ("หมายเหตุ",          "TRACE_NOTE",        "555555", 30),
    ]

    # ── Header row ───────────────────────────────────
    ws.row_dimensions[1].height = 36
    for ci, (hdr_txt, _, color, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr_txt)
        apply(c, **hdr_style(color))
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)
        ].width = width

    # ── Data rows ───────────────────────────────────
    STATUS_FILL = {
        "OK":        PatternFill("solid", fgColor="E2EFDA"),
        "NO_METER":  PatternFill("solid", fgColor="FCE4D6"),
        "NO_TX":     PatternFill("solid", fgColor="FFF2CC"),
        "NO_FEEDER": PatternFill("solid", fgColor="FFF2CC"),
    }

    for ri, row_data in enumerate(rows, 2):
        status     = row_data.get("TRACE_STATUS", "")
        row_fill   = STATUS_FILL.get(status)
        ws.row_dimensions[ri].height = 18

        for ci, (_, field_key, _, _) in enumerate(COLS, 1):
            val = row_data.get(field_key)
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font, c.border, c.alignment = BODY, BORDER, WRAP
            if row_fill:
                c.fill = row_fill

    ws.freeze_panes = "A2"

    # ── Summary sheet ────────────────────────────────
    ws2 = wb.create_sheet("สรุป")
    total     = len(rows)
    ok        = sum(1 for r in rows if r.get("TRACE_STATUS") == "OK")
    no_meter  = sum(1 for r in rows if r.get("TRACE_STATUS") == "NO_METER")
    no_tx     = sum(1 for r in rows if r.get("TRACE_STATUS") == "NO_TX")
    no_feeder = sum(1 for r in rows if r.get("TRACE_STATUS") == "NO_FEEDER")
    with_rc   = sum(1 for r in rows if (r.get("RC_COUNT") or 0) > 0)
    with_cb   = sum(1 for r in rows if (r.get("CB_COUNT") or 0) > 0)

    summary = [
        ("วันที่รัน",                  datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Meter ทั้งหมด",              total),
        ("Trace สำเร็จ (OK)",          ok),
        ("ไม่พบ Meter ใน Layer 26",    no_meter),
        ("ไม่พบ Transformer",          no_tx),
        ("ไม่สามารถระบุ Feeder",       no_feeder),
        ("มี Recloser ≥ 1 ตัว",        with_rc),
        ("มี Circuit Breaker ≥ 1 ตัว", with_cb),
    ]
    for ri, (label, val) in enumerate(summary, 1):
        c_lbl = ws2.cell(row=ri, column=1, value=label)
        c_lbl.font   = Font(name="Arial", size=10, bold=True)
        c_lbl.border = BORDER
        c_val = ws2.cell(row=ri, column=2, value=val)
        c_val.font   = Font(name="Arial", size=10)
        c_val.border = BORDER
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22

    wb.save(path)
    print(f"\n✅ บันทึก → {path}")

# ─────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("PEA GIS Upstream Trace v2.0")
    print(f"เริ่มต้น: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # ── โหลด Excel input ──────────────────────────────
    try:
        import openpyxl
    except ImportError:
        print("กรุณารัน: pip install openpyxl")
        return

    print(f"\nโหลด {INPUT_FILE} ...")
    try:
        wb  = openpyxl.load_workbook(INPUT_FILE, read_only=True)
        ws  = wb.active
        raw = list(ws.iter_rows(values_only=True))
    except FileNotFoundError:
        print(f"❌ ไม่พบไฟล์ {INPUT_FILE}")
        return

    # คอลัมน์ B = "หมายเลขเครื่องวัด PEA"
    pea_numbers = [
        str(row[1]).strip()
        for row in raw[1:]                      # ข้าม header
        if row[1] is not None and str(row[1]).strip()
    ]
    print(f"พบ {len(pea_numbers)} หมายเลข PEA")

    # ── ดึง Switch domain ─────────────────────────────
    print("\nขั้นตอน 0: ดึง Switch/Fuse subtype domain ...")
    subtype_map = load_switch_subtypes()

    # ── วนลูป Trace ───────────────────────────────────
    results = []
    total   = len(pea_numbers)

    print(f"\nเริ่ม Trace {total} Meter ...")
    print("-" * 60)

    for i, peano in enumerate(pea_numbers, 1):
        print(f"[{i:3d}/{total}] {peano}", end=" ")
        t0  = time.time()
        res = trace_upstream(peano, subtype_map)
        elapsed = time.time() - t0
        results.append(res)

        status = res["TRACE_STATUS"]
        tx_id  = res.get("TX_FACILITYID") or "-"
        feeder = res.get("METER_FEEDERID") or "-"
        rc_n   = res.get("RC_COUNT", 0)
        sw_n   = res.get("SW_COUNT", 0)
        cb_n   = res.get("CB_COUNT", 0)

        print(f"| {status:10s} | TX:{tx_id:12s} | Feeder:{feeder:8s}"
              f" | RC:{rc_n} SW:{sw_n} CB:{cb_n} | {elapsed:.1f}s")

    # ── บันทึกผล ──────────────────────────────────────
    print("\n" + "=" * 60)
    ok_n = sum(1 for r in results if r["TRACE_STATUS"] == "OK")
    print(f"สรุป: OK={ok_n}  ไม่สำเร็จ={total - ok_n}  จากทั้งหมด {total}")
    save_results(results, OUTPUT_FILE)
    print(f"สิ้นสุด: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
