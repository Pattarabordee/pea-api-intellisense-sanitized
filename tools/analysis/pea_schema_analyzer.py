"""
PEA GIS — Schema Analyzer
วิเคราะห์โครงสร้าง (Schema) ของ Layer ที่เกี่ยวข้องกับ Upstream Trace

รันบนเครื่องที่เชื่อมต่ออินเทอร์เน็ต (หรืออยู่ใน Intranet PEA สำหรับ URL ภายใน)

Dependencies:
    pip install openpyxl
    (ไม่ต้องติดตั้ง requests — ใช้ urllib มาตรฐาน)

Usage:
    python pea_schema_analyzer.py
    
Output:
    - พิมพ์ผลลัพธ์ใน Terminal
    - สร้างไฟล์  pea_schema_report.xlsx
"""

import urllib.request
import urllib.parse
import json
import time
import sys
from datetime import datetime

# ─────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────
BASE_EXT = "https://gisne1.pea.co.th/arcgis/rest/services"
BASE_INT  = "http://172.16.184.233/arcgis/rest/services"   # ใช้ได้เฉพาะใน Intranet PEA

# Layer ที่ต้องวิเคราะห์
LAYERS = [
    {
        "service":   "PEA",
        "layer_id":  26,
        "alias":     "DS_LowVoltageMeter",
        "desc":      "มิเตอร์แรงต่ำ (จุดเริ่มต้น Trace)",
        "join_note": "field ที่ต้องมีเพื่อ join: FEEDERID, PEANO/FACILITYID",
    },
    {
        "service":   "PEA",
        "layer_id":  17,
        "alias":     "DS_Transformer",
        "desc":      "หม้อแปลงไฟฟ้า",
        "join_note": "ต้องมี FEEDERID ที่รูปแบบเดียวกับ Layer 26",
    },
    {
        "service":   "PEA",
        "layer_id":  14,
        "alias":     "DS_Recloser",
        "desc":      "รีโคลสเซอร์ (อุปกรณ์ป้องกัน HV)",
        "join_note": "ต้องมี FEEDERID, OPERATIONTYPE, LOCATION",
    },
    {
        "service":   "PEA",
        "layer_id":  16,
        "alias":     "DS_Switch",
        "desc":      "สวิตซ์/ฟิวส์แรงสูง",
        "join_note": "ต้องมี field แยกประเภท Fuse vs Switch",
    },
    {
        "service":   "PEA",
        "layer_id":  11,
        "alias":     "DS_CircuitBreaker",
        "desc":      "เซอร์กิตเบรกเกอร์ 22-33 kV",
        "join_note": "อุปกรณ์ป้องกันสูงสุดใน Feeder",
    },
    {
        "service":   "PEA_QUERY",
        "layer_id":  29,
        "alias":     "DS_Meter (PEA_QUERY)",
        "desc":      "มิเตอร์พร้อมชื่อลูกค้า",
        "join_note": "มี PEANO, CUSTOMERNAME, FEEDERID",
    },
]

# Field สำคัญที่อยากรู้ว่ามีหรือไม่
KEY_FIELDS = [
    "FEEDERID", "FACILITYID", "PEANO", "OBJECTID", "GLOBALID",
    "OPERATIONTYPE", "RATEDKVA", "LOCATION", "CUSTOMERNAME",
    "ACCOUNTNUMBER", "INSTALLATIONID", "PHASECODE",
    "CIRCUITID", "SUBSTATIONID",
]

# ─────────────────────────────────────────
#  HTTP HELPER
# ─────────────────────────────────────────
def fetch(url: str, timeout: int = 30, retries: int = 3) -> dict:
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "PEA-Schema-Analyzer/1.0"})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                raw = r.read().decode("utf-8", errors="replace")
                return json.loads(raw)
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)   # exponential back-off
            else:
                return {"__error__": str(e)}

# ─────────────────────────────────────────
#  SCHEMA FETCH
# ─────────────────────────────────────────
def get_layer_info(service: str, layer_id: int, base: str = BASE_EXT) -> dict:
    url = f"{base}/{service}/MapServer/{layer_id}?f=pjson"
    return fetch(url)

def get_sample_rows(service: str, layer_id: int, n: int = 3, base: str = BASE_EXT) -> dict:
    params = urllib.parse.urlencode({
        "where":             "1=1",
        "outFields":         "*",
        "returnGeometry":    "false",
        "resultRecordCount": n,
        "f":                 "pjson",
    })
    url = f"{base}/{service}/MapServer/{layer_id}/query?{params}"
    return fetch(url)

def get_count(service: str, layer_id: int, base: str = BASE_EXT) -> int:
    params = urllib.parse.urlencode({
        "where":           "1=1",
        "returnCountOnly": "true",
        "f":               "pjson",
    })
    url = f"{base}/{service}/MapServer/{layer_id}/query?{params}"
    data = fetch(url)
    return data.get("count", -1)

def get_feederid_sample(service: str, layer_id: int, base: str = BASE_EXT) -> list:
    """ดึงตัวอย่างค่า FEEDERID ที่ไม่ซ้ำ 10 ค่า"""
    params = urllib.parse.urlencode({
        "where":                "FEEDERID IS NOT NULL",
        "outFields":            "FEEDERID",
        "returnGeometry":       "false",
        "returnDistinctValues": "true",
        "orderByFields":        "FEEDERID",
        "resultRecordCount":    10,
        "f":                    "pjson",
    })
    url = f"{base}/{service}/MapServer/{layer_id}/query?{params}"
    data = fetch(url)
    feats = data.get("features", [])
    return [f["attributes"].get("FEEDERID") for f in feats if f.get("attributes")]

# ─────────────────────────────────────────
#  ANALYSIS
# ─────────────────────────────────────────
def analyze_layer(lyr: dict, base: str = BASE_EXT) -> dict:
    svc = lyr["service"]
    lid = lyr["layer_id"]

    print(f"\n{'='*60}")
    print(f"Layer {lid}: {lyr['alias']}  ({svc})")
    print(f"  {lyr['desc']}")
    print(f"{'='*60}")

    # 1. Layer Info
    print("  [1/4] Layer Info ...")
    info = get_layer_info(svc, lid, base)
    time.sleep(0.4)

    if "__error__" in info:
        print(f"  ERROR: {info['__error__']}")
        return {"meta": lyr, "error": info["__error__"]}

    fields = info.get("fields", [])
    geom_type = info.get("geometryType", "N/A")
    max_rec   = info.get("maxRecordCount", "N/A")
    has_query = "Query" in (info.get("capabilities") or "")

    print(f"  geometryType  : {geom_type}")
    print(f"  maxRecordCount: {max_rec}")
    print(f"  capabilities  : {info.get('capabilities', 'N/A')}")
    print(f"  จำนวน fields  : {len(fields)}")

    # 2. Key Fields Check
    print("\n  [2/4] Key Fields ...")
    field_names = [f["name"].upper() for f in fields]
    field_map   = {f["name"].upper(): f for f in fields}
    found, missing = [], []
    for kf in KEY_FIELDS:
        if kf.upper() in field_names:
            f_info = field_map[kf.upper()]
            found.append({
                "name":   f_info["name"],
                "type":   f_info.get("type", ""),
                "length": f_info.get("length", ""),
                "alias":  f_info.get("alias", ""),
            })
            print(f"    ✅ {kf:20s} | {f_info.get('type','')} | len={f_info.get('length','')}")
        else:
            missing.append(kf)
            print(f"    ❌ {kf:20s} | ไม่มีใน Layer นี้")

    # 3. Row Count
    print("\n  [3/4] Row Count ...")
    count = get_count(svc, lid, base)
    time.sleep(0.4)
    print(f"  จำนวน records ทั้งหมด: {count:,}" if count >= 0 else "  ดึง count ไม่ได้")

    # 4. FEEDERID Sample
    feederid_samples = []
    if "FEEDERID" in field_names:
        print("\n  [4/4] ตัวอย่าง FEEDERID ...")
        feederid_samples = get_feederid_sample(svc, lid, base)
        time.sleep(0.4)
        for fid in feederid_samples:
            print(f"    → {fid}")
    else:
        print("\n  [4/4] ไม่มี FEEDERID ข้ามขั้นตอนนี้")

    # 5. Sample Row (1 แถว)
    print("\n  [5/5] Sample Row ...")
    sample = get_sample_rows(svc, lid, n=1, base=base)
    time.sleep(0.4)
    sample_attrs = {}
    feats = sample.get("features", [])
    if feats:
        sample_attrs = feats[0].get("attributes", {})
        for k, v in sample_attrs.items():
            print(f"    {k:30s}: {v}")
    else:
        print(f"  ไม่มี sample: {sample.get('__error__', 'features=[]')}")

    return {
        "meta":             lyr,
        "geom_type":        geom_type,
        "max_record_count": max_rec,
        "total_count":      count,
        "all_fields":       fields,
        "key_found":        found,
        "key_missing":      missing,
        "feederid_samples": feederid_samples,
        "sample_row":       sample_attrs,
    }

# ─────────────────────────────────────────
#  CROSS-LAYER FEEDERID FORMAT CHECK
# ─────────────────────────────────────────
def compare_feederid(all_results: list):
    print(f"\n{'='*60}")
    print("CROSS-LAYER: เปรียบเทียบรูปแบบ FEEDERID")
    print(f"{'='*60}")
    for r in all_results:
        if "error" in r:
            continue
        samples = r.get("feederid_samples", [])
        alias = r["meta"]["alias"]
        lid   = r["meta"]["layer_id"]
        if samples:
            print(f"  Layer {lid} ({alias}): {samples[:5]}")
        else:
            print(f"  Layer {lid} ({alias}): ไม่มี FEEDERID")

# ─────────────────────────────────────────
#  EXCEL REPORT
# ─────────────────────────────────────────
def save_excel_report(all_results: list, path: str = "pea_schema_report.xlsx"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n⚠️  ติดตั้ง openpyxl ก่อน: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ─ Style helpers ─
    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    SUB_FILL   = PatternFill("solid", fgColor="BDD7EE")
    SUB_FONT   = Font(bold=True, name="Arial", size=10)
    OK_FILL    = PatternFill("solid", fgColor="C6EFCE")
    NO_FILL    = PatternFill("solid", fgColor="FFC7CE")
    BODY_FONT  = Font(name="Arial", size=9)
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
    thin       = Side(style="thin", color="B8B8B8")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def sub(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill, c.font, c.border = SUB_FILL, SUB_FONT, BORDER
        return c

    def cell(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font, c.border, c.alignment = BODY_FONT, BORDER, WRAP_ALIGN
        if fill:
            c.fill = fill
        return c

    # ════════════════════════════════
    #  Sheet 1: Layer Summary
    # ════════════════════════════════
    ws1 = wb.active
    ws1.title = "Layer Summary"

    COLS = ["Layer ID", "Service", "Layer Name", "คำอธิบาย",
            "Geometry", "maxRecordCount", "จำนวน records",
            "Key Fields ที่มี", "Key Fields ที่ขาด",
            "ตัวอย่าง FEEDERID"]
    for ci, col in enumerate(COLS, 1):
        hdr(ws1, 1, ci, col)
        ws1.row_dimensions[1].height = 30

    for ri, r in enumerate(all_results, 2):
        m = r["meta"]
        if "error" in r:
            cell(ws1, ri, 1, m["layer_id"])
            cell(ws1, ri, 2, m["service"])
            cell(ws1, ri, 3, m["alias"])
            cell(ws1, ri, 4, f"ERROR: {r['error']}", fill=NO_FILL)
            continue

        found_names   = [f["name"] for f in r.get("key_found", [])]
        missing_names = r.get("key_missing", [])
        fid_samples   = ", ".join(str(x) for x in r.get("feederid_samples", []))
        cnt           = r.get("total_count", -1)

        cell(ws1, ri, 1, m["layer_id"])
        cell(ws1, ri, 2, m["service"])
        cell(ws1, ri, 3, m["alias"])
        cell(ws1, ri, 4, m["desc"])
        cell(ws1, ri, 5, r.get("geom_type", ""))
        cell(ws1, ri, 6, r.get("max_record_count", ""))
        cell(ws1, ri, 7, cnt if cnt >= 0 else "N/A")
        cell(ws1, ri, 8, ", ".join(found_names),
             fill=OK_FILL if found_names else None)
        cell(ws1, ri, 9, ", ".join(missing_names),
             fill=NO_FILL if missing_names else OK_FILL)
        cell(ws1, ri, 10, fid_samples)

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 28
    ws1.column_dimensions["E"].width = 18
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 14
    ws1.column_dimensions["H"].width = 40
    ws1.column_dimensions["I"].width = 40
    ws1.column_dimensions["J"].width = 35
    ws1.freeze_panes = "A2"

    # ════════════════════════════════
    #  Sheet 2: All Fields (ทุก field ของทุก Layer)
    # ════════════════════════════════
    ws2 = wb.create_sheet("All Fields")
    FCOLS = ["Layer ID", "Layer Name", "Field Name", "Alias",
             "Type", "Length", "เป็น Key Field?"]
    for ci, col in enumerate(FCOLS, 1):
        hdr(ws2, 1, ci, col)
    ws2.row_dimensions[1].height = 28

    row = 2
    for r in all_results:
        if "error" in r:
            continue
        m  = r["meta"]
        kf = {f["name"].upper() for f in r.get("key_found", [])}
        for f in r.get("all_fields", []):
            is_key = "✅" if f["name"].upper() in kf else ""
            fill   = OK_FILL if is_key else None
            cell(ws2, row, 1, m["layer_id"])
            cell(ws2, row, 2, m["alias"])
            cell(ws2, row, 3, f.get("name", ""))
            cell(ws2, row, 4, f.get("alias", ""))
            cell(ws2, row, 5, f.get("type", ""))
            cell(ws2, row, 6, f.get("length", ""))
            cell(ws2, row, 7, is_key, fill=fill)
            row += 1

    for col, w in zip("ABCDEFG", [8, 22, 22, 28, 24, 8, 12]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # ════════════════════════════════
    #  Sheet 3: Sample Rows
    # ════════════════════════════════
    ws3 = wb.create_sheet("Sample Rows")
    sub(ws3, 1, 1, "Layer")
    sub(ws3, 1, 2, "Field")
    sub(ws3, 1, 3, "Sample Value")

    row = 2
    for r in all_results:
        if "error" in r or not r.get("sample_row"):
            continue
        m = r["meta"]
        for k, v in r["sample_row"].items():
            cell(ws3, row, 1, f"[{m['layer_id']}] {m['alias']}")
            cell(ws3, row, 2, k)
            cell(ws3, row, 3, str(v) if v is not None else "NULL")
            row += 1
        # separator
        row += 1

    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 40
    ws3.freeze_panes = "A2"

    # ════════════════════════════════
    #  Sheet 4: FEEDERID Comparison
    # ════════════════════════════════
    ws4 = wb.create_sheet("FEEDERID Comparison")
    hdr(ws4, 1, 1, "Layer")
    hdr(ws4, 1, 2, "ตัวอย่าง FEEDERID (10 ค่าแรก)")
    hdr(ws4, 1, 3, "หมายเหตุ")

    row = 2
    for r in all_results:
        if "error" in r:
            continue
        m = r["meta"]
        samples = r.get("feederid_samples", [])
        cell(ws4, row, 1, f"[{m['layer_id']}] {m['alias']}")
        cell(ws4, row, 2, "  |  ".join(str(s) for s in samples) if samples else "ไม่มี FEEDERID")
        cell(ws4, row, 3, m.get("join_note", ""))
        row += 1

    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 60
    ws4.column_dimensions["C"].width = 45

    # ─ Save ─
    wb.save(path)
    print(f"\n✅ บันทึก Excel report → {path}")

# ─────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────
def main():
    print("=" * 60)
    print("PEA GIS Schema Analyzer")
    print(f"เวลาเริ่มต้น: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    print(f"\nBase URL: {BASE_EXT}")
    print("(เปลี่ยนเป็น BASE_INT ถ้าอยู่ใน Intranet PEA)")

    all_results = []
    for lyr in LAYERS:
        result = analyze_layer(lyr, base=BASE_EXT)
        all_results.append(result)

    compare_feederid(all_results)

    # สรุปสิ่งที่ต้องสังเกต
    print(f"\n{'='*60}")
    print("สรุปสิ่งที่ต้องตรวจสอบเพิ่มเติม")
    print(f"{'='*60}")
    for r in all_results:
        if "error" in r:
            print(f"\n  ⛔  Layer {r['meta']['layer_id']} {r['meta']['alias']}: {r['error']}")
            continue
        missing = r.get("key_missing", [])
        if missing:
            print(f"\n  ⚠️   Layer {r['meta']['layer_id']} {r['meta']['alias']}")
            print(f"      ไม่มี field: {', '.join(missing)}")
            print(f"      → อาจต้องหาทางอื่นสำหรับ join")

    save_excel_report(all_results, "pea_schema_report.xlsx")

    print(f"\nเวลาสิ้นสุด: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == "__main__":
    main()
