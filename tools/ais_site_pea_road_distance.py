from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
from html import unescape
import json
import math
from pathlib import Path
import re
import sqlite3
import statistics
import sys
import time
import urllib.error
import urllib.parse
import urllib.request

from openpyxl import load_workbook


PEA_OFFICE_SEARCH_URL = "https://sbiz.pea.co.th/pea-office/search.php"
OSRM_TABLE_URL = "https://router.project-osrm.org/table/v1/driving"
DEFAULT_WORKBOOK = Path("Meter_ID_NE For PEA_12เดือน_LatLong_R01 1.xlsx")
DEFAULT_OUTPUT = Path("runtime/ais_site_nearest_pea_office_road_distance.csv")
DEFAULT_OFFICE_OUTPUT = Path("runtime/pea_office_coordinate_sources.csv")
DEFAULT_SUMMARY = Path("runtime/ais_site_nearest_pea_office_road_distance_summary.md")
DEFAULT_SQLITE = Path("runtime/ais_site_pea_road_distance.sqlite")
DEFAULT_OFFICIAL_CACHE = Path("runtime/pea_office_search_cache.html")
DEFAULT_COORD_CACHE = Path("runtime/pea_office_coordinate_cache.json")
DEFAULT_OSRM_CACHE = Path("runtime/osrm_nearest_pea_office_cache.json")


OFFICE_REQUEST_TEXT = [
    "กฟจ.อุดรธานี (L)",
    "กฟจ.หนองคาย (L)",
    "กฟจ.ขอนแก่น (L)",
    "กฟจ.เลย (M)",
    "กฟจ.สกลนคร (L)",
    "กฟจ.นครพนม (L)",
    "กฟส.ชุมแพ (M)",
    "กฟส.กุมภวาปี (M)",
    "กฟส.หนองหาน (M)",
    "กฟส.พังโคน (M)",
    "กฟจ.หนองบัวลำภู (M)",
    "กฟส.บ้านไผ่ (M)",
    "กฟจ.บึงกาฬ (M)",
    "กฟส.เมืองขอนแก่น 2 (L)",
    "กฟส.เมืองอุดรธานี 2 (L)",
    "กฟส.สว่างแดนดิน (M)",
    "กฟส.วังสะพุง (M)",
    "กฟจ.อุบลราชธานี (L)",
    "กฟจ.ศรีสะเกษ (S)",
    "กฟจ.ยโสธร (S)",
    "กฟจ.มหาสารคาม (S)",
    "กฟจ.กาฬสินธุ์ (S)",
    "กฟจ.ร้อยเอ็ด (S)",
    "กฟจ.มุกดาหาร (S)",
    "กฟจ.อำนาจเจริญ (S)",
    "กฟส.เสลภูมิ (M)",
    "กฟส.สมเด็จ (M)",
    "กฟส.กันทรลักษ์ (L)",
    "กฟส.เดชอุดม (L)",
    "กฟส.วารินชำราบ (M)",
    "กฟส.ตระการพืชผล (M)",
    "กฟส.กันทรารมย์ (M)",
    "กฟส.กันทรวิชัย (M)",
    "กฟจ.นครราชสีมา",
    "กฟจ.ชัยภูมิ",
    "กฟจ.บุรีรัมย์",
    "กฟจ.สุรินทร์",
    "กฟส.นางรอง",
    "กฟส.ปากช่อง",
    "กฟส.สีคิ้ว",
    "กฟส.บัวใหญ่",
    "กฟส.โชคชัย",
    "กฟส.พิมาย",
    "กฟส.ภูเขียว",
    "กฟส.ปราสาท",
    "กฟส.เมืองนครราชสีมา2(หัวทะเล)",
    "กฟส.เมืองนครราชสีมา3(สุรนารี)",
    "กฟส.ปักธงชัย",
]


COORDINATE_SOURCES = {
    "กฟจ.อุดรธานี (L)": {"kind": "longdo", "id": "A00017172"},
    "กฟจ.หนองคาย (L)": {"kind": "longdo", "id": "A10289415"},
    "กฟจ.ขอนแก่น (L)": {"kind": "longdo", "id": "A10289687"},
    "กฟจ.เลย (M)": {"kind": "longdo", "id": "A10289455"},
    "กฟจ.สกลนคร (L)": {"kind": "longdo", "id": "A10289609"},
    "กฟจ.นครพนม (L)": {"kind": "longdo", "id": "A10289388"},
    "กฟส.ชุมแพ (M)": {"kind": "longdo", "id": "A10289403"},
    "กฟส.กุมภวาปี (M)": {"kind": "longdo", "id": "A10289619"},
    "กฟส.หนองหาน (M)": {"kind": "longdo", "id": "A10289496"},
    "กฟส.พังโคน (M)": {"kind": "longdo", "id": "A10289389"},
    "กฟจ.หนองบัวลำภู (M)": {"kind": "longdo", "id": "A10289287"},
    "กฟส.บ้านไผ่ (M)": {"kind": "longdo", "id": "A10289494"},
    "กฟจ.บึงกาฬ (M)": {"kind": "longdo", "id": "A10289236"},
    "กฟส.เมืองขอนแก่น 2 (L)": {"kind": "longdo", "id": "A10289244"},
    "กฟส.เมืองอุดรธานี 2 (L)": {"kind": "longdo", "id": "A10289586"},
    "กฟส.สว่างแดนดิน (M)": {"kind": "longdo", "id": "A10289056"},
    "กฟส.วังสะพุง (M)": {"kind": "longdo", "id": "A10289171"},
    "กฟจ.อุบลราชธานี (L)": {"kind": "longdo", "id": "A10289350"},
    "กฟจ.ศรีสะเกษ (S)": {"kind": "longdo", "id": "A10289666"},
    "กฟจ.ยโสธร (S)": {"kind": "longdo", "id": "A10289627"},
    "กฟจ.มหาสารคาม (S)": {"kind": "longdo", "id": "A10289675"},
    "กฟจ.กาฬสินธุ์ (S)": {"kind": "longdo", "id": "A10289847"},
    "กฟจ.ร้อยเอ็ด (S)": {"kind": "longdo", "id": "A10289592"},
    "กฟจ.มุกดาหาร (S)": {"kind": "longdo", "id": "A10289761"},
    "กฟจ.อำนาจเจริญ (S)": {"kind": "longdo", "id": "A10289828"},
    "กฟส.เสลภูมิ (M)": {"kind": "longdo", "id": "A10289471"},
    "กฟส.สมเด็จ (M)": {"kind": "longdo", "id": "A10289775"},
    "กฟส.กันทรลักษ์ (L)": {"kind": "longdo", "id": "A10289783"},
    "กฟส.เดชอุดม (L)": {"kind": "longdo", "id": "A10289453"},
    "กฟส.วารินชำราบ (M)": {
        "kind": "contact_page",
        "url": "https://xn--th-th862016----6l6b0pjdawau1b4g4czke4d0h7m.contact.page/map",
    },
    "กฟส.ตระการพืชผล (M)": {"kind": "longdo", "id": "A10289313"},
    "กฟส.กันทรารมย์ (M)": {"kind": "longdo", "id": "A10289073"},
    "กฟส.กันทรวิชัย (M)": {"kind": "longdo", "id": "A10289474"},
    "กฟจ.นครราชสีมา": {"kind": "longdo", "id": "A10583481"},
    "กฟจ.ชัยภูมิ": {"kind": "longdo", "id": "A10289655"},
    "กฟจ.บุรีรัมย์": {"kind": "longdo", "id": "A10289408"},
    "กฟจ.สุรินทร์": {"kind": "longdo", "id": "A10289685"},
    "กฟส.นางรอง": {"kind": "longdo", "id": "A10289719"},
    "กฟส.ปากช่อง": {"kind": "longdo", "id": "A10289832"},
    "กฟส.สีคิ้ว": {"kind": "longdo", "id": "A10289893"},
    "กฟส.บัวใหญ่": {"kind": "longdo", "id": "A10289822"},
    "กฟส.โชคชัย": {"kind": "longdo", "id": "A10289435"},
    "กฟส.พิมาย": {"kind": "longdo", "id": "A10289750"},
    "กฟส.ภูเขียว": {"kind": "longdo", "id": "A10289741"},
    "กฟส.ปราสาท": {"kind": "longdo", "id": "A10289527"},
    "กฟส.เมืองนครราชสีมา2(หัวทะเล)": {"kind": "longdo", "id": "A10289500"},
    "กฟส.เมืองนครราชสีมา3(สุรนารี)": {"kind": "longdo", "id": "A10289506"},
    "กฟส.ปักธงชัย": {"kind": "longdo", "id": "A10289362"},
}


OFFICIAL_BRANCH_ALIASES = {
    "เมืองขอนแก่น 2": ["ขอนแก่น 2"],
    "เมืองอุดรธานี 2": ["อุดรธานี 2"],
    "เมืองนครราชสีมา2(หัวทะเล)": ["นครราชสีมา 2 (หัวทะเล)"],
    "เมืองนครราชสีมา3(สุรนารี)": ["นครราชสีมา 3 (สุรนารี)"],
}


@dataclass(frozen=True)
class OfficeRequest:
    office_id: str
    requested_office: str
    abbr: str
    branch: str
    size: str


@dataclass(frozen=True)
class OfficialOfficeRecord:
    name: str
    address: str
    phone: str
    province_class: str


@dataclass(frozen=True)
class Office:
    office_id: str
    requested_office: str
    size: str
    branch: str
    official_name: str
    official_address: str
    phone: str
    lat: float
    lon: float
    coordinate_source: str
    coordinate_url: str
    coordinate_status: str


@dataclass(frozen=True)
class Site:
    site_ref: str
    site_code: str
    company: str
    source_province: str
    lat: float
    lon: float


def normalize_thai(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ")
    return re.sub(r"[\s\.\-_/()（）]+", "", text.strip())


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_requested_offices(texts: list[str] = OFFICE_REQUEST_TEXT) -> list[OfficeRequest]:
    requests: list[OfficeRequest] = []
    for idx, text in enumerate(texts, start=1):
        match = re.match(r"^(กฟ[จส])\.(.+)$", text.strip())
        if not match:
            raise ValueError(f"Unsupported PEA office request: {text}")
        abbr = match.group(1)
        branch_with_size = match.group(2).strip()
        size = ""
        size_match = re.search(r"\(([LMS])\)\s*$", branch_with_size)
        if size_match:
            size = size_match.group(1)
            branch = branch_with_size[: size_match.start()].strip()
        else:
            branch = branch_with_size.strip()
        requests.append(
            OfficeRequest(
                office_id=f"pea_{idx:02d}",
                requested_office=text.strip(),
                abbr=abbr,
                branch=branch,
                size=size,
            )
        )
    return requests


def strip_tags(fragment: str) -> str:
    without_tags = re.sub(r"<[^>]+>", "", fragment)
    return clean_text(unescape(without_tags))


def parse_pea_office_records(html: str) -> list[OfficialOfficeRecord]:
    panel_re = re.compile(
        r'<div class="col-md-6 province (?P<province_class>[^"]+)">.*?'
        r'<div class="panel-body">(?P<body>.*?)</div></div></div>',
        re.S,
    )
    paragraph_re = re.compile(r"<p>(.*?)</p>", re.S)
    records: list[OfficialOfficeRecord] = []
    for panel in panel_re.finditer(html):
        parts = [strip_tags(part) for part in paragraph_re.findall(panel.group("body"))]
        parts = [part for part in parts if part]
        if not parts:
            continue
        phone = ""
        address_parts = []
        for part in parts[1:]:
            if part.startswith("โทร"):
                phone = part
            else:
                address_parts.append(part)
        records.append(
            OfficialOfficeRecord(
                name=parts[0],
                address=clean_text(" ".join(address_parts)),
                phone=phone,
                province_class=panel.group("province_class"),
            )
        )
    return records


def office_match_score(request: OfficeRequest, record: OfficialOfficeRecord) -> int:
    name = normalize_thai(record.name)
    branch_variants = [request.branch, *OFFICIAL_BRANCH_ALIASES.get(request.branch, [])]
    if request.abbr == "กฟจ":
        exact = normalize_thai(f"การไฟฟ้าส่วนภูมิภาคจังหวัด{request.branch}")
        if name == exact:
            return 120
        if "เขต" in name:
            return -1
        if normalize_thai(f"จังหวัด{request.branch}") in name:
            return 100
        return -1

    for variant_index, branch_variant in enumerate(branch_variants):
        branch = normalize_thai(branch_variant)
        phrases = [
            normalize_thai(f"สาขา{branch_variant}"),
            normalize_thai(f"อำเภอ{branch_variant}"),
            normalize_thai(f"จังหวัด{branch_variant}"),
        ]
        for offset, phrase in enumerate(phrases):
            if phrase in name:
                return 110 - offset - variant_index
        if branch and branch in name:
            return 50 - variant_index
    return -1


def match_official_record(
    request: OfficeRequest, records: list[OfficialOfficeRecord]
) -> tuple[OfficialOfficeRecord | None, int]:
    scored = [
        (office_match_score(request, record), record)
        for record in records
    ]
    scored = [(score, record) for score, record in scored if score >= 0]
    if not scored:
        return None, -1
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1], scored[0][0]


def fetch_text(url: str, timeout: int = 45) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 AIS-PEA-road-distance/1.0",
            "Accept": "text/html,application/json;q=0.9,*/*;q=0.8",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        raw = response.read()
    return raw.decode("utf-8", errors="replace")


def load_official_records(cache_path: Path, refresh: bool = False) -> list[OfficialOfficeRecord]:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if cache_path.exists() and not refresh:
        html = cache_path.read_text(encoding="utf-8", errors="replace")
    else:
        html = fetch_text(PEA_OFFICE_SEARCH_URL)
        cache_path.write_text(html, encoding="utf-8")
    return parse_pea_office_records(html)


def longdo_info_url(longdo_id: str) -> str:
    return f"https://map.longdo.com/main/p/{longdo_id}/info"


def extract_longdo_lat_lon(html: str) -> tuple[float, float] | None:
    match = re.search(r"snippet/\?lat=([-0-9.]+)&(?:long|lon)=([-0-9.]+)", html)
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def extract_contact_page_lat_lon(html: str) -> tuple[float, float] | None:
    lat_match = re.search(r'data-lat="([-0-9.]+)"', html)
    lon_match = re.search(r'data-(?:long|lon)="([-0-9.]+)"', html)
    if not lat_match or not lon_match:
        return None
    return float(lat_match.group(1)), float(lon_match.group(1))


def load_json(path: Path, default: object) -> object:
    if not path.exists():
        return default
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)


def source_url(source: dict[str, str]) -> str:
    if source["kind"] == "longdo":
        return longdo_info_url(source["id"])
    return source["url"]


def fetch_coordinate(
    source: dict[str, str],
    cache: dict[str, dict[str, object]],
    refresh: bool = False,
) -> tuple[float, float, str]:
    url = source_url(source)
    if url in cache and not refresh:
        entry = cache[url]
        return float(entry["lat"]), float(entry["lon"]), str(entry.get("status", "cached"))

    html = fetch_text(url)
    if source["kind"] == "longdo":
        coord = extract_longdo_lat_lon(html)
    elif source["kind"] == "contact_page":
        coord = extract_contact_page_lat_lon(html)
    else:
        raise ValueError(f"Unsupported coordinate source: {source['kind']}")
    if coord is None:
        raise ValueError(f"Could not extract coordinate from {url}")

    lat, lon = coord
    cache[url] = {
        "lat": lat,
        "lon": lon,
        "status": "fetched",
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "url": url,
    }
    return lat, lon, "fetched"


def build_offices(
    official_records: list[OfficialOfficeRecord],
    coord_cache_path: Path,
    refresh_coordinates: bool = False,
) -> list[Office]:
    coord_cache = load_json(coord_cache_path, {})
    if not isinstance(coord_cache, dict):
        coord_cache = {}

    offices: list[Office] = []
    missing_sources = []
    for request in parse_requested_offices():
        source = COORDINATE_SOURCES.get(request.requested_office)
        if source is None:
            missing_sources.append(request.requested_office)
            continue
        official, _score = match_official_record(request, official_records)
        lat, lon, coordinate_status = fetch_coordinate(source, coord_cache, refresh_coordinates)
        offices.append(
            Office(
                office_id=request.office_id,
                requested_office=request.requested_office,
                size=request.size,
                branch=request.branch,
                official_name=official.name if official else "",
                official_address=official.address if official else "",
                phone=official.phone if official else "",
                lat=lat,
                lon=lon,
                coordinate_source=source["kind"],
                coordinate_url=source_url(source),
                coordinate_status=coordinate_status,
            )
        )
    if missing_sources:
        raise ValueError(f"Missing coordinate source for: {', '.join(missing_sources)}")
    write_json(coord_cache_path, coord_cache)
    return offices


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def load_sites(workbook_path: Path, limit: int | None = None) -> tuple[list[Site], dict[str, int]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["Joined"]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    header = {name: idx for idx, name in enumerate(headers)}
    required = ["Location ID", "Province", "Com", "LAT", "LONG"]
    missing = [name for name in required if name not in header]
    if missing:
        raise ValueError(f"Missing required workbook columns: {', '.join(missing)}")

    sites: list[Site] = []
    seen_keys: set[tuple[str, float, float]] = set()
    ref_counts: dict[str, int] = {}
    stats = {
        "workbook_rows": 0,
        "invalid_coordinates": 0,
        "duplicate_site_coordinate": 0,
    }

    for row in rows:
        stats["workbook_rows"] += 1
        lat = parse_float(row[header["LAT"]])
        lon = parse_float(row[header["LONG"]])
        if lat is None or lon is None or not (5 <= lat <= 21 and 97 <= lon <= 106):
            stats["invalid_coordinates"] += 1
            continue

        location_id = clean_text(row[header["Location ID"]])
        site_code = clean_text(row[header["SITE Code"]]) if "SITE Code" in header else ""
        base_ref = location_id or site_code or f"row_{stats['workbook_rows']}"
        key = (base_ref, round(lat, 7), round(lon, 7))
        if key in seen_keys:
            stats["duplicate_site_coordinate"] += 1
            continue
        seen_keys.add(key)
        ref_counts[base_ref] = ref_counts.get(base_ref, 0) + 1
        site_ref = base_ref if ref_counts[base_ref] == 1 else f"{base_ref}__coord{ref_counts[base_ref]}"

        sites.append(
            Site(
                site_ref=site_ref,
                site_code=site_code,
                company=clean_text(row[header["Com"]]),
                source_province=clean_text(row[header["Province"]]),
                lat=lat,
                lon=lon,
            )
        )
        if limit is not None and len(sites) >= limit:
            break
    wb.close()
    stats["sites_loaded"] = len(sites)
    return sites, stats


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius_km = 6371.0088
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)
    a = (
        math.sin(d_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    )
    return 2 * radius_km * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def select_nearest_office(
    distances_m: list[float | None],
    durations_s: list[float | None],
    site: Site,
    offices: list[Office],
) -> dict[str, object]:
    best_idx = None
    best_distance = None
    for idx, distance in enumerate(distances_m):
        if distance is None:
            continue
        if best_distance is None or distance < best_distance:
            best_idx = idx
            best_distance = distance

    if best_idx is None or best_distance is None:
        straight = [
            haversine_km(site.lat, site.lon, office.lat, office.lon)
            for office in offices
        ]
        best_idx = min(range(len(offices)), key=lambda idx: straight[idx])
        office = offices[best_idx]
        return {
            "nearest_office_id": office.office_id,
            "road_distance_km": "",
            "osrm_duration_min": "",
            "straight_line_km": round(straight[best_idx], 3),
            "route_status": "osrm_no_route",
        }

    office = offices[best_idx]
    duration = durations_s[best_idx] if durations_s and durations_s[best_idx] is not None else None
    return {
        "nearest_office_id": office.office_id,
        "road_distance_km": round(best_distance / 1000, 3),
        "osrm_duration_min": round(duration / 60, 1) if duration is not None else "",
        "straight_line_km": round(haversine_km(site.lat, site.lon, office.lat, office.lon), 3),
        "route_status": "ok",
    }


def osrm_cache_signature(offices: list[Office]) -> str:
    material = "|".join(
        f"{office.office_id}:{office.lat:.7f},{office.lon:.7f}" for office in offices
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def site_cache_key(site: Site) -> str:
    return f"{site.site_ref}|{site.lat:.7f},{site.lon:.7f}"


def osrm_table_batch(
    sites: list[Site],
    offices: list[Office],
    retries: int = 4,
    timeout: int = 90,
) -> tuple[list[list[float | None]], list[list[float | None]]]:
    coords = [f"{site.lon:.7f},{site.lat:.7f}" for site in sites]
    coords.extend(f"{office.lon:.7f},{office.lat:.7f}" for office in offices)
    source_indexes = ";".join(str(idx) for idx in range(len(sites)))
    destination_indexes = ";".join(str(len(sites) + idx) for idx in range(len(offices)))
    url = f"{OSRM_TABLE_URL}/{';'.join(coords)}?sources={source_indexes}&destinations={destination_indexes}&annotations=distance,duration"

    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            text = fetch_text(url, timeout=timeout)
            payload = json.loads(text)
            if payload.get("code") != "Ok":
                raise RuntimeError(f"OSRM returned {payload.get('code')}: {payload.get('message')}")
            return payload["distances"], payload.get("durations", [])
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, RuntimeError) as exc:
            last_error = exc
            if attempt == retries:
                break
            time.sleep(min(20, 2**attempt))
    raise RuntimeError(f"OSRM table request failed after {retries} attempts: {last_error}")


def compute_nearest_by_road(
    sites: list[Site],
    offices: list[Office],
    batch_size: int,
    osrm_cache_path: Path,
    refresh_osrm: bool = False,
    sleep_seconds: float = 0.2,
) -> list[dict[str, object]]:
    signature = osrm_cache_signature(offices)
    cache = load_json(osrm_cache_path, {"office_signature": signature, "items": {}})
    if not isinstance(cache, dict) or cache.get("office_signature") != signature or refresh_osrm:
        cache = {"office_signature": signature, "items": {}}
    cache.setdefault("items", {})
    items = cache["items"]
    if not isinstance(items, dict):
        items = {}
        cache["items"] = items

    results_by_key: dict[str, dict[str, object]] = {}
    missing_sites: list[Site] = []
    for site in sites:
        key = site_cache_key(site)
        if key in items:
            results_by_key[key] = dict(items[key])
        else:
            missing_sites.append(site)

    total_batches = math.ceil(len(missing_sites) / batch_size) if missing_sites else 0
    for batch_number, start in enumerate(range(0, len(missing_sites), batch_size), start=1):
        batch = missing_sites[start : start + batch_size]
        distances, durations = osrm_table_batch(batch, offices)
        for idx, site in enumerate(batch):
            row = select_nearest_office(distances[idx], durations[idx] if durations else [], site, offices)
            key = site_cache_key(site)
            items[key] = row
            results_by_key[key] = row
        write_json(osrm_cache_path, cache)
        processed = min(start + len(batch), len(missing_sites))
        print(
            f"OSRM batch {batch_number}/{total_batches}: {processed}/{len(missing_sites)} missing sites routed",
            flush=True,
        )
        if sleep_seconds:
            time.sleep(sleep_seconds)

    rows = []
    office_by_id = {office.office_id: office for office in offices}
    for site in sites:
        nearest = results_by_key[site_cache_key(site)]
        office = office_by_id[str(nearest["nearest_office_id"])]
        row = {
            "site_ref": site.site_ref,
            "site_code": site.site_code,
            "company": site.company,
            "source_province": site.source_province,
            "site_lat": round(site.lat, 7),
            "site_lon": round(site.lon, 7),
            "nearest_pea_office": office.requested_office,
            "nearest_office_size": office.size,
            "official_office_name": office.official_name,
            "office_address": office.official_address,
            "office_lat": round(office.lat, 7),
            "office_lon": round(office.lon, 7),
            "road_distance_km": nearest["road_distance_km"],
            "straight_line_km": nearest["straight_line_km"],
            "osrm_duration_min": nearest["osrm_duration_min"],
            "route_status": nearest["route_status"],
            "office_coord_source": office.coordinate_source,
            "office_coord_url": office.coordinate_url,
        }
        rows.append(row)
    return rows


def write_csv_rows(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_office_sources(path: Path, offices: list[Office]) -> None:
    fieldnames = [
        "office_id",
        "requested_office",
        "size",
        "official_office_name",
        "official_address",
        "phone",
        "office_lat",
        "office_lon",
        "coordinate_source",
        "coordinate_url",
        "coordinate_status",
    ]
    rows = [
        {
            "office_id": office.office_id,
            "requested_office": office.requested_office,
            "size": office.size,
            "official_office_name": office.official_name,
            "official_address": office.official_address,
            "phone": office.phone,
            "office_lat": round(office.lat, 7),
            "office_lon": round(office.lon, 7),
            "coordinate_source": office.coordinate_source,
            "coordinate_url": office.coordinate_url,
            "coordinate_status": office.coordinate_status,
        }
        for office in offices
    ]
    write_csv_rows(path, rows, fieldnames)


def write_sqlite(path: Path, rows: list[dict[str, object]], offices: list[Office], metadata: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as conn:
        conn.execute("DROP TABLE IF EXISTS nearest_site_office")
        conn.execute("DROP TABLE IF EXISTS pea_office_coordinates")
        conn.execute("DROP TABLE IF EXISTS run_metadata")
        conn.execute(
            """
            CREATE TABLE nearest_site_office (
                site_ref TEXT,
                site_code TEXT,
                company TEXT,
                source_province TEXT,
                site_lat REAL,
                site_lon REAL,
                nearest_pea_office TEXT,
                nearest_office_size TEXT,
                official_office_name TEXT,
                office_address TEXT,
                office_lat REAL,
                office_lon REAL,
                road_distance_km REAL,
                straight_line_km REAL,
                osrm_duration_min REAL,
                route_status TEXT,
                office_coord_source TEXT,
                office_coord_url TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE pea_office_coordinates (
                office_id TEXT,
                requested_office TEXT,
                size TEXT,
                official_office_name TEXT,
                official_address TEXT,
                phone TEXT,
                office_lat REAL,
                office_lon REAL,
                coordinate_source TEXT,
                coordinate_url TEXT,
                coordinate_status TEXT
            )
            """
        )
        conn.execute("CREATE TABLE run_metadata (key TEXT PRIMARY KEY, value TEXT)")
        if rows:
            conn.executemany(
                f"INSERT INTO nearest_site_office VALUES ({','.join(['?'] * len(rows[0]))})",
                [[row[field] if row[field] != "" else None for field in rows[0].keys()] for row in rows],
            )
        office_rows = [
            [
                office.office_id,
                office.requested_office,
                office.size,
                office.official_name,
                office.official_address,
                office.phone,
                office.lat,
                office.lon,
                office.coordinate_source,
                office.coordinate_url,
                office.coordinate_status,
            ]
            for office in offices
        ]
        conn.executemany(
            "INSERT INTO pea_office_coordinates VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            office_rows,
        )
        conn.executemany(
            "INSERT INTO run_metadata VALUES (?,?)",
            [(str(key), json.dumps(value, ensure_ascii=False)) for key, value in metadata.items()],
        )
        conn.commit()


def write_summary(
    path: Path,
    rows: list[dict[str, object]],
    offices: list[Office],
    site_stats: dict[str, int],
    metadata: dict[str, object],
) -> None:
    road_values = [
        float(row["road_distance_km"])
        for row in rows
        if row["road_distance_km"] not in ("", None)
    ]
    counts: dict[str, int] = {}
    for row in rows:
        counts[str(row["nearest_pea_office"])] = counts.get(str(row["nearest_pea_office"]), 0) + 1
    top_counts = sorted(counts.items(), key=lambda item: item[1], reverse=True)[:15]
    route_status_counts: dict[str, int] = {}
    for row in rows:
        status = str(row["route_status"])
        route_status_counts[status] = route_status_counts.get(status, 0) + 1

    lines = [
        "# AIS site to nearest PEA office road distance",
        "",
        f"- Generated: {metadata['generated_at']}",
        f"- AIS site source: `{metadata['site_workbook']}` sheet `Joined`",
        f"- PEA official office source: {PEA_OFFICE_SEARCH_URL}",
        f"- Office coordinate sources: Longdo public POI pages plus one contact.page fallback for Warin Chamrap",
        f"- Road route source: OSRM table API `{OSRM_TABLE_URL}`",
        f"- Offices considered: {len(offices)}",
        f"- Site rows scanned: {site_stats['workbook_rows']}",
        f"- Site coordinates output: {len(rows)}",
        f"- Invalid coordinate rows skipped: {site_stats['invalid_coordinates']}",
        f"- Duplicate site-coordinate rows skipped: {site_stats['duplicate_site_coordinate']}",
        "",
        "## Distance summary",
        "",
    ]
    if road_values:
        lines.extend(
            [
                f"- Min road distance km: {min(road_values):.3f}",
                f"- Median road distance km: {statistics.median(road_values):.3f}",
                f"- Mean road distance km: {statistics.mean(road_values):.3f}",
                f"- Max road distance km: {max(road_values):.3f}",
            ]
        )
    else:
        lines.append("- No OSRM road distances were returned.")

    lines.extend(["", "## Route status", ""])
    for status, count in sorted(route_status_counts.items()):
        lines.append(f"- {status}: {count}")

    lines.extend(["", "## Top nearest offices by site count", "", "| Office | Site count |", "|---|---:|"])
    for office, count in top_counts:
        lines.append(f"| {office} | {count} |")

    lines.extend(
        [
            "",
            "## Redaction",
            "",
            "Output excludes meter numbers, CA values, PEANO lists, tokens, raw WebEx text, and customer identity fields.",
            "",
            "## Output files",
            "",
            f"- CSV: `{metadata['output_csv']}`",
            f"- Office source CSV: `{metadata['office_csv']}`",
            f"- SQLite: `{metadata['sqlite']}`",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(args: argparse.Namespace) -> dict[str, object]:
    workbook = Path(args.workbook)
    official_records = load_official_records(Path(args.official_cache), args.refresh_official)
    offices = build_offices(official_records, Path(args.coord_cache), args.refresh_coordinates)
    sites, site_stats = load_sites(workbook, args.limit_sites)
    if not sites:
        raise RuntimeError("No valid AIS site coordinates loaded.")

    rows = compute_nearest_by_road(
        sites,
        offices,
        batch_size=args.batch_size,
        osrm_cache_path=Path(args.osrm_cache),
        refresh_osrm=args.refresh_osrm,
        sleep_seconds=args.sleep_seconds,
    )

    fieldnames = list(rows[0].keys())
    output = Path(args.output)
    office_output = Path(args.office_output)
    sqlite_output = Path(args.sqlite)
    summary_output = Path(args.summary)

    metadata = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "site_workbook": str(workbook),
        "output_csv": str(output),
        "office_csv": str(office_output),
        "sqlite": str(sqlite_output),
        "summary": str(summary_output),
        "site_stats": site_stats,
        "osrm_batch_size": args.batch_size,
    }

    write_csv_rows(output, rows, fieldnames)
    write_office_sources(office_output, offices)
    write_sqlite(sqlite_output, rows, offices, metadata)
    write_summary(summary_output, rows, offices, site_stats, metadata)
    return {
        "rows": len(rows),
        "offices": len(offices),
        "output": str(output),
        "office_output": str(office_output),
        "summary": str(summary_output),
        "sqlite": str(sqlite_output),
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compute road distance from AIS cell sites to the nearest requested PEA office."
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--office-output", default=str(DEFAULT_OFFICE_OUTPUT))
    parser.add_argument("--summary", default=str(DEFAULT_SUMMARY))
    parser.add_argument("--sqlite", default=str(DEFAULT_SQLITE))
    parser.add_argument("--official-cache", default=str(DEFAULT_OFFICIAL_CACHE))
    parser.add_argument("--coord-cache", default=str(DEFAULT_COORD_CACHE))
    parser.add_argument("--osrm-cache", default=str(DEFAULT_OSRM_CACHE))
    parser.add_argument("--batch-size", type=int, default=40)
    parser.add_argument("--sleep-seconds", type=float, default=0.2)
    parser.add_argument("--limit-sites", type=int)
    parser.add_argument("--refresh-official", action="store_true")
    parser.add_argument("--refresh-coordinates", action="store_true")
    parser.add_argument("--refresh-osrm", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    if args.batch_size < 1:
        parser.error("--batch-size must be >= 1")
    if args.batch_size + len(OFFICE_REQUEST_TEXT) > 100:
        parser.error("--batch-size plus office count must be <= 100 for public OSRM safety")
    result = run(args)
    print(json.dumps(result, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
